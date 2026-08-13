from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import get_user_model
from .forms import RegistroPacienteForm, CitaForm, CitaEnfermeraForm, ControlPrenatalForm, EditarPacienteEnfermeraForm
from .models import Usuario, LogAuditoria
from citas.models import Cita
from datetime import time, timedelta
from django.db import transaction, connection
from django.db.models import Q, Count
from control_prenatal.models import ControlPrenatal
from django.http import JsonResponse, HttpResponse
from django.contrib.sessions.models import Session
from django.utils import timezone
from django.contrib.auth import update_session_auth_hash
from django.urls import reverse
from django.core.paginator import Paginator
import csv
import json
import logging

logger = logging.getLogger(__name__)



# ── HELPERS AUDITORÍA ──────────────────────────────────────────

def get_limite_cancelacion():
    """Retorna (hoy, hora_limite) con 30 min de margen para auto-cancelar."""
    ahora_dt = timezone.localtime()
    limite_dt = ahora_dt - timedelta(minutes=30)
    return limite_dt.date(), limite_dt.time()


def auto_cancelar_citas(queryset):
    """Cancela citas pendientes cuya fecha+hora ya superó los 30 min de margen."""
    hoy, limite = get_limite_cancelacion()
    queryset.filter(estado='pendiente').filter(
        Q(fecha__lt=hoy) |
        Q(fecha=hoy, hora__lt=limite)
    ).update(estado='cancelada')


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def registrar_log(request, accion, modulo='', descripcion='', severidad='INFO'):
    """Registra una acción en el log de auditoría."""
    try:
        LogAuditoria.objects.create(
            usuario     = request.user if request.user.is_authenticated else None,
            accion      = accion,
            modulo      = modulo,
            descripcion = descripcion,
            ip_address  = get_client_ip(request),
            severidad   = severidad,
        )
    except Exception:
        pass  # Nunca romper el flujo por un log fallido


def q_usuarios_con_acceso_prenatal():
    return Q(paciente__estado_embarazo='ACTIVO') | Q(citas_paciente__medico__medico__especialidad__tipo='prenatal')


def q_pacientes_con_acceso_prenatal():
    return Q(estado_embarazo='ACTIVO') | Q(usuario__citas_paciente__medico__medico__especialidad__tipo='prenatal')


def q_citas_con_acceso_prenatal():
    return Q(paciente__paciente__estado_embarazo='ACTIVO') | Q(medico__medico__especialidad__tipo='prenatal')


def es_genero_femenino(valor):
    """Verifica si el género es femenino. Si está vacío/None, se asume femenino por compatibilidad."""
    if not valor or valor.strip() == '':
        # Si el género no está definido, asumimos femenino para pacientes legacy
        return True
    return (valor or '').strip().lower() in ('femenino', 'f')


User = get_user_model()
 
from functools import wraps


def no_cache_view(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        response = view_func(request, *args, **kwargs)
        response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response
    return wrapper


def paciente_prenatal_required(view_func):
    """
    Decorator: solo pacientes con módulo prenatal activo pueden acceder.
    Acepta:
      - tipo_paciente == 'prenatal'  (cuentas prenatales puras)
      - tipo_paciente == 'general' con tiene_prenatal == True (cuentas generales con módulo activado)
    """
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        rol_lower = request.user.rol.lower() if request.user.rol else ''
        if rol_lower != 'paciente':
            return redireccionar_por_rol(request.user)
        if not request.user.puede_prenatal:
            # Es paciente general sin módulo prenatal → su dashboard general
            return redirect('paciente_general_dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


def medico_required(view_func):
    """Decorator: solo médicos (cualquier especialidad) pueden acceder."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        rol_lower = request.user.rol.lower() if request.user.rol else ''
        if rol_lower != 'medico':
            return redireccionar_por_rol(request.user)
        return view_func(request, *args, **kwargs)
    return wrapper


def medico_prenatal_required(view_func):
    """Decorator: solo médicos pueden acceder."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        rol_lower = request.user.rol.lower() if request.user.rol else ''
        if rol_lower != 'medico':
            return redireccionar_por_rol(request.user)
        # Permitir acceso a todos los médicos
        return view_func(request, *args, **kwargs)
    return wrapper


def admin_required(view_func):
    """Decorator: solo administradores pueden acceder."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        rol_lower = request.user.rol.lower() if request.user.rol else ''
        if rol_lower != 'admin':
            return redireccionar_por_rol(request.user)
        return view_func(request, *args, **kwargs)
    return wrapper


def enfermera_required(view_func):
    """Decorator: solo enfermeras y secretarias pueden acceder."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        rol_lower = request.user.rol.lower() if request.user.rol else ''
        if rol_lower not in ['enfermera', 'secretaria']:
            return redireccionar_por_rol(request.user)
        return view_func(request, *args, **kwargs)
    return wrapper

@no_cache_view
def login_view(request):
    # Si ya está autenticado, redirigir directo a su panel
    if request.user.is_authenticated:
        return redireccionar_por_rol(request.user)

    especialidad_id = request.GET.get('especialidad') or request.POST.get('especialidad_id', '')
    tipo_paciente = request.GET.get('tipo') or request.POST.get('tipo_paciente') or 'general'

    if request.method == 'POST':
        username_or_email = request.POST.get('username', '').strip()
        password = request.POST.get('password')
        tipo_esperado = request.POST.get('tipo_paciente') or 'general'

        # Buscar por username O por email
        user_obj = None
        try:
            user_obj = User.objects.get(username=username_or_email)
        except User.DoesNotExist:
            # Intentar buscar por email
            try:
                user_obj = User.objects.get(email__iexact=username_or_email)
            except User.DoesNotExist:
                pass
            except User.MultipleObjectsReturned:
                messages.error(request, 'Hay más de una cuenta con ese correo. Ingresa con tu usuario.')
                return render(request, 'login.html', {'especialidad_id': especialidad_id, 'tipo_paciente': tipo_paciente})

        if user_obj:
            # Verificar si la cuenta está desactivada
            if user_obj.check_password(password) and not user_obj.is_active:
                registrar_log(request, 'ERROR', 'Autenticación',
                    f'Intento de acceso a cuenta desactivada: "{user_obj.username}"', 'WARNING')
                messages.error(request, 'CUENTA_DESACTIVADA')
                return render(request, 'login.html', {'especialidad_id': especialidad_id, 'tipo_paciente': tipo_paciente})
            # Autenticar usando el username real del usuario encontrado
            user = authenticate(request, username=user_obj.username, password=password)
        else:
            user = None

        if user is not None:
            login(request, user)
            registrar_log(request, 'LOGIN', 'Autenticación',
                f'Inicio de sesión exitoso — rol: {user.rol}', 'INFO')
            
            # Si es paciente, check tipo esperado
            if user.rol == 'paciente':
                if tipo_esperado == 'prenatal' and not user.puede_prenatal:
                    messages.info(request, 'Tu cuenta no tiene el módulo prenatal activo. Solicita a la enfermera que lo active.')
            
            return redireccionar_por_rol(user)
        else:
            # Detectar si fue cuenta desactivada o credenciales incorrectas
            if user_obj and user_obj.check_password(password) and not user_obj.is_active:
                pass  # ya manejado arriba
            else:
                registrar_log(request, 'ERROR', 'Autenticación',
                    f'Intento de login fallido para usuario: "{username_or_email}"', 'WARNING')
            messages.error(request, 'Usuario/correo o contraseña incorrectos')

    return render(request, 'login.html', {'especialidad_id': especialidad_id, 'tipo_paciente': tipo_paciente})
 
 
@no_cache_view
def logout_view(request):

    rol = getattr(request.user, 'rol', None)

    tipo_paciente = getattr(request.user, 'tipo_paciente', None)

    # Guardar nombre ANTES de logout, porque después request.user es AnonymousUser

    nombre_usuario = ''

    if request.user.is_authenticated:

        nombre_usuario = request.user.get_full_name() or request.user.username

    registrar_log(request, 'LOGOUT', 'Autenticación',

        f'Cierre de sesión — usuario: {nombre_usuario}', 'INFO')

    logout(request)
    return redirect('login')
 
 
def redireccionar_por_rol(user):
    rol_lower = user.rol.lower() if user.rol else ''
    
    if rol_lower == 'admin':
        return redirect('admin_dashboard')
    elif rol_lower == 'medico':
        return redirect('medico_dashboard')
    elif rol_lower in ['enfermera', 'secretaria']:
        return redirect('enfermera_dashboard')
    elif rol_lower == 'paciente':
        # Si es general pura, va al dashboard general
        # Si tiene módulo prenatal activo (sea general o prenatal), va al dashboard prenatal
        if user.puede_prenatal:
            return redirect('paciente_dashboard')
        else:
            return redirect('paciente_general_dashboard')
    else:
        return redirect('login')
 
 
@login_required
@no_cache_view
def admin_dashboard(request):
    rol_lower = request.user.rol.lower() if request.user.rol else ''
    if rol_lower != 'admin':
        return redireccionar_por_rol(request.user)

    from pacientes.models import Paciente
    from django.utils import timezone

    hoy = timezone.localdate()
    ahora = timezone.localtime().time()

    # ── AUTO-CANCELAR citas pendientes vencidas (30 min de margen) ──
    auto_cancelar_citas(Cita.objects.filter(estado='pendiente'))

    total_usuarios = User.objects.count()
    total_medicos = User.objects.filter(rol='medico').count()
    total_pacientes = Paciente.objects.count()
    total_pacientes_generales = User.objects.filter(rol='paciente', paciente__estado_embarazo='NINGUNO').count()
    total_pacientes_prenatales = User.objects.filter(
        rol='paciente', paciente__estado_embarazo='ACTIVO'
    ).distinct().count()
    total_citas = Cita.objects.count()
    citas_hoy = Cita.objects.filter(fecha=hoy).count()
    citas_pendientes = Cita.objects.filter(estado='pendiente').count()
    citas_atendidas = Cita.objects.filter(estado='atendido').count()
    citas_canceladas = Cita.objects.filter(estado='cancelado').count()
    total_controles = ControlPrenatal.objects.count()
    citas_recientes = Cita.objects.select_related('paciente', 'medico').order_by('-fecha', 'hora')[:8]

    return render(request, 'admin/dashboard_admin.html', {
        'total_usuarios': total_usuarios,
        'total_medicos': total_medicos,
        'total_pacientes': total_pacientes,
        'total_pacientes_generales': total_pacientes_generales,
        'total_pacientes_prenatales': total_pacientes_prenatales,
        'total_citas': total_citas,
        'citas_hoy': citas_hoy,
        'citas_pendientes': citas_pendientes,
        'citas_atendidas': citas_atendidas,
        'citas_canceladas': citas_canceladas,
        'total_controles': total_controles,
        'citas_recientes': citas_recientes,
        'hoy': hoy,
    })
 

@login_required
@no_cache_view
def medico_dashboard(request):
    rol_lower = request.user.rol.lower() if request.user.rol else ''
    if rol_lower != 'medico':
        return redireccionar_por_rol(request.user)

    from pacientes.models import Paciente
    from citas.models import Cita
    from control_prenatal.models import ControlPrenatal
    from django.db.models import Q

    # Determinar si es médico prenatal o general
    try:
        es_prenatal = request.user.medico.especialidad and \
                      request.user.medico.especialidad.tipo == 'prenatal'
    except:
        es_prenatal = True

    # ── AUTO-CANCELAR citas pendientes (30 min de margen) ──
    auto_cancelar_citas(Cita.objects.filter(medico=request.user))

    # ── STATS ──
    hoy = timezone.now().date()
    total_citas               = Cita.objects.filter(medico=request.user).count()
    citas_pendientes          = Cita.objects.filter(medico=request.user, estado='pendiente', fecha=hoy).count()
    citas_pendientes_futuras  = Cita.objects.filter(medico=request.user, estado='pendiente').count()
    citas_atendidas           = Cita.objects.filter(medico=request.user, estado='atendido').count()
    citas_canceladas          = Cita.objects.filter(medico=request.user, estado__in=['cancelado', 'cancelada']).count()
    proximas_citas            = Cita.objects.filter(medico=request.user, estado='pendiente').order_by('fecha', 'hora')[:5]

    # ── PACIENTES: solo del tipo que corresponde al médico ──
    if es_prenatal:
        total_pacientes = Usuario.objects.filter(
            rol='paciente', paciente__estado_embarazo='ACTIVO'
        ).distinct().count()
    else:
        total_pacientes = Usuario.objects.filter(
            rol='paciente',
            paciente__estado_embarazo='NINGUNO'
        ).count()

    context = {
        'total_citas':              total_citas,
        'citas_pendientes':         citas_pendientes,
        'citas_pendientes_futuras': citas_pendientes_futuras,
        'citas_atendidas':          citas_atendidas,
        'citas_canceladas':         citas_canceladas,
        'proximas_citas':           proximas_citas,
        'total_pacientes':          total_pacientes,
        'es_prenatal':              es_prenatal,
    }

    if es_prenatal:
        total_controles = ControlPrenatal.objects.filter(medico=request.user).count()
        context['total_controles'] = total_controles
        return render(request, 'medico/dashboard_medico.html', context)
    else:
        return render(request, 'medico/dashboard_medico_general.html', context)
 
 
@login_required
@login_required
@no_cache_view
def enfermera_dashboard(request):
    rol_lower = request.user.rol.lower() if request.user.rol else ''
    if rol_lower not in ['enfermera', 'secretaria']:
        return redireccionar_por_rol(request.user)

    # ── AUTO-CANCELAR citas pendientes (30 min de margen) ──
    auto_cancelar_citas(Cita.objects)
    hoy = timezone.now().date()

    from pacientes.models import Paciente
    return render(request, 'enfermera/dashboard_enfermera.html', {
        'total_citas':      Cita.objects.count(),
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
        'citas_atendidas':  Cita.objects.filter(estado='atendido').count(),
        'total_pacientes':  Paciente.objects.count(),
        'citas_recientes':  Cita.objects.order_by('-fecha', 'hora')[:5],
    })
 
 
@paciente_prenatal_required
@no_cache_view
def paciente_dashboard(request):
    # paciente_prenatal_required ya valida rol y tipo_paciente
    
    try:
        print(f"DEBUG: Usuario accediendo al dashboard: {request.user.username}")

        # ── AUTO-CANCELAR citas pendientes (30 min de margen) ──
        auto_cancelar_citas(Cita.objects.filter(paciente=request.user))
        hoy = timezone.now().date()

        from pacientes.models import Paciente
        from paciente_general.models import ProgramacionParto
        try:
            perfil = Paciente.objects.get(usuario=request.user)
            print(f"DEBUG: Perfil encontrado: {perfil}")
        except Paciente.DoesNotExist:
            perfil = None
            print("DEBUG: No se encontró perfil de paciente")

        # ── Citas y estadísticas ─────────────────────────────────────────────
        proxima_cita = Cita.objects.filter(
            paciente=request.user, fecha__gte=hoy, estado__in=['pendiente', 'confirmada']
        ).order_by('fecha', 'hora').first()
        print(f"DEBUG: Próxima cita: {proxima_cita}")
        
        total_citas = Cita.objects.filter(paciente=request.user).count()
        print(f"DEBUG: Total citas: {total_citas}")
        
        citas_pendientes = Cita.objects.filter(
            paciente=request.user, estado__in=['pendiente', 'confirmada'], fecha__gte=hoy
        ).count()
        print(f"DEBUG: Citas pendientes: {citas_pendientes}")
        
        ultimas_citas = Cita.objects.filter(paciente=request.user).order_by('-fecha', 'hora')[:3]
        print(f"DEBUG: Últimas citas: {ultimas_citas.count()}")

        # ── Programación de parto próxima ─────────────────────────────────────
        tiene_prenatal_activo = bool(
            perfil
            and getattr(perfil, 'tiene_prenatal', False)
            and getattr(perfil, 'estado_embarazo', '') == 'ACTIVO'
        )
        parto_programado = None
        if tiene_prenatal_activo:
            print(f"DEBUG: Buscando programación de parto para usuario {request.user.id} - {request.user.username}")
            programaciones = ProgramacionParto.objects.filter(paciente=request.user)
            print(f"DEBUG: Total programaciones encontradas: {programaciones.count()}")
            for prog in programaciones:
                print(f"DEBUG: Programación - ID:{prog.id}, Estado:{prog.estado}, Fecha:{prog.fecha_programada}")
            
            parto_programado = ProgramacionParto.objects.filter(
                paciente=request.user,
                estado__in=['programado', 'confirmado'],
            ).order_by('fecha_programada', 'hora_programada').first()
            print(f"DEBUG: Parto programado encontrado: {parto_programado}")
        else:
            print(f"DEBUG: No tiene prenatal activo - perfil:{perfil}, tiene_prenatal:{getattr(perfil, 'tiene_prenatal', False) if perfil else None}, estado:{getattr(perfil, 'estado_embarazo', '') if perfil else None}")
        print(f"DEBUG: Parto programado: {parto_programado}")

        print("DEBUG: Renderizando template...")
        response = render(request, 'paciente/dashboard_paciente.html', {
            'user': request.user,
            'perfil': perfil,
            'proxima_cita': proxima_cita,
            'total_citas': total_citas,
            'citas_pendientes': citas_pendientes,
            'ultimas_citas': ultimas_citas,
            'hoy': hoy,
            'parto_programado': parto_programado,
        })
        print("DEBUG: Template renderizado exitosamente")
        response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        return response
    except Exception as e:
        print(f"ERROR EN paciente_dashboard: {e}")
        import traceback
        traceback.print_exc()
        raise
 
def registro_paciente(request):
    # ?tipo=ginecologia, ?tipo=medicina, etc. (podemos recibir la especialidad)
    especialidad_param = request.GET.get('tipo') or request.POST.get('especialidad_param') or ''
    tipo = request.GET.get('tipo_paciente') or request.POST.get('tipo_paciente') or 'general'

    if request.method == 'POST':
        post_data = request.POST.copy()
        if tipo == 'prenatal':
            post_data['genero'] = 'femenino'
        form = RegistroPacienteForm(post_data)
        if form.is_valid():
            try:
                user = form.save(commit=True)
                if tipo == 'prenatal':
                    from pacientes.models import Paciente
                    paciente = Paciente.objects.get(usuario=user)
                    _aplicar_tipo_paciente(paciente, 'prenatal', user)
                login(request, user)
                request.session['mensaje_bienvenida_registro'] = f'¡Cuenta creada con éxito! Bienvenido/a, {user.first_name}.'
                return redirect('paciente_general_dashboard')
            except Exception as e:
                messages.error(request, f'Error al crear la cuenta: {str(e)}')
        else:
            # Recopilar todos los errores del form en mensajes legibles
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, error)
                    else:
                        label = form.fields[field].label if field in form.fields else field
                        messages.error(request, f'{label}: {error}')
    else:
        form = RegistroPacienteForm()

    return render(request, 'registro.html', {
        'form': form,
        'especialidad_param': especialidad_param,
    })


def verificar_disponibilidad(request):
    """
    Endpoint AJAX (GET) — verifica si username, cédula o email ya existen.
    ?campo=username&valor=XXX  |  ?campo=cedula&valor=XXX  |  ?campo=email&valor=XXX
    """
    campo = request.GET.get('campo', '').strip()
    valor = request.GET.get('valor', '').strip()

    if not campo or not valor:
        return JsonResponse({'disponible': False, 'mensaje': 'Datos incompletos.'})

    if campo == 'username':
        existe = Usuario.objects.filter(username__iexact=valor).exists()
        if existe:
            return JsonResponse({'disponible': False, 'mensaje': 'Este usuario ya está registrado.'})
        return JsonResponse({'disponible': True, 'mensaje': 'Usuario disponible.'})

    if campo == 'cedula':
        from pacientes.models import Paciente
        existe = Paciente.objects.filter(cedula=valor).exists()
        if existe:
            return JsonResponse({'disponible': False, 'mensaje': 'Esta cédula ya está registrada.'})
        return JsonResponse({'disponible': True, 'mensaje': 'Cédula disponible.'})

    if campo == 'email':
        existe = Usuario.objects.filter(email__iexact=valor).exists()
        if existe:
            return JsonResponse({'disponible': False, 'mensaje': 'Este correo ya está registrado.'})
        return JsonResponse({'disponible': True, 'mensaje': 'Correo disponible.'})

    return JsonResponse({'disponible': False, 'mensaje': 'Campo no válido.'})


@paciente_prenatal_required
@no_cache_view
def mi_perfil(request):
    # paciente_prenatal_required ya valida rol y tipo_paciente

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'perfil':
            email = request.POST.get('email', '').strip()
            if email and Usuario.objects.filter(email__iexact=email).exclude(id=request.user.id).exists():
                messages.error(request, 'Este correo ya está registrado.')
                return redirect('mi_perfil')
            request.user.first_name = request.POST.get('first_name', '')
            request.user.last_name = request.POST.get('last_name', '')
            request.user.email = email
            request.user.save()
            registrar_log(request, 'UPDATE', 'Perfil',
                'Paciente prenatal actualizó sus datos personales', 'INFO')
            messages.success(request, 'Datos actualizados correctamente.')

        elif action == 'password':
            from django.contrib.auth import update_session_auth_hash
            password_actual = request.POST.get('password_actual')
            password_nuevo = request.POST.get('password_nuevo')
            password_confirmar = request.POST.get('password_confirmar')

            if not request.user.check_password(password_actual):
                messages.error(request, 'La contraseña actual es incorrecta.')
            elif password_nuevo != password_confirmar:
                messages.error(request, 'Las contraseñas nuevas no coinciden.')
            elif len(password_nuevo) < 8:
                messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
            else:
                request.user.set_password(password_nuevo)
                request.user.save()
                update_session_auth_hash(request, request.user)
                registrar_log(request, 'UPDATE', 'Perfil',
                    'Paciente prenatal cambió su contraseña', 'INFO')
                messages.success(request, 'Contraseña actualizada correctamente.')

        return redirect('mi_perfil')

    return render(request, 'paciente/mi_perfil.html')
 
 
@paciente_prenatal_required
@no_cache_view
def agendar_cita(request):
    from datetime import datetime
    # paciente_prenatal_required ya valida rol y tipo_paciente

    from medicos.models import Medico
    from landing.models import Especialidad

    # Todas las especialidades (excepto partos) y todos los médicos activos
    # IMPORTANTE: Permitir CUALQUIER especialidad, no solo prenatal
    especialidades_prenatales = Especialidad.objects.filter(activo=True).exclude(nombre__icontains='parto').exclude(nombre__icontains='cesárea')
    medicos_prenatales = Medico.objects.filter(
        usuario__is_active=True
    ).select_related('usuario', 'especialidad')

    if request.method == 'POST':
        medico_id  = request.POST.get('medico_id')
        fecha      = request.POST.get('fecha')
        hora       = request.POST.get('hora')
        motivo     = request.POST.get('motivo', '').strip()

        errores = []
        if not medico_id:
            errores.append('Debes seleccionar un médico.')
        if not fecha:
            errores.append('Debes seleccionar una fecha.')
        if not hora:
            errores.append('Debes seleccionar una hora.')

        if not errores:
            try:
                medico_usuario = Usuario.objects.get(id=medico_id, rol='medico')
                existe = Cita.objects.filter(
                    medico=medico_usuario, fecha=fecha, hora=hora,
                    estado__in=['pendiente', 'confirmada']
                ).exists()
                if existe:
                    errores.append('Esta hora ya está ocupada para el médico seleccionado. Elige otra hora.')
                else:
                    cita = Cita(
                        paciente=request.user,
                        medico=medico_usuario,
                        fecha=fecha,
                        hora=hora,
                        motivo=motivo,
                    )
                    # Asignar especialidad del médico automáticamente
                    try:
                        perfil_medico = Medico.objects.get(usuario=medico_usuario)
                        if perfil_medico.especialidad:
                            cita.especialidad = perfil_medico.especialidad
                    except Medico.DoesNotExist:
                        pass
                    cita.save()
                    registrar_log(request, 'CREATE', 'Citas',
                        f'Cita agendada por {request.user.get_full_name()} '
                        f'para el {fecha} a las {hora}')
                    messages.success(request, 'Cita agendada correctamente.')
                    return redirect('ver_citas')
            except Usuario.DoesNotExist:
                errores.append('El médico seleccionado no es válido.')

        for e in errores:
            messages.error(request, e)

    # Horas disponibles default (se actualizan via AJAX según día seleccionado)
    horas_semana = [h.strftime("%H:%M") for h in HORAS_DISPONIBLES_SEMANA]
    horas_finde = [h.strftime("%H:%M") for h in HORAS_DISPONIBLES_FINDE]

    return render(request, 'paciente/agendar_cita.html', {
        'medicos': medicos_prenatales,
        'especialidades': especialidades_prenatales,
        'horas_semana': horas_semana,
        'horas_finde': horas_finde,
    })
 
@paciente_prenatal_required
@no_cache_view
def ver_citas(request):
    # paciente_prenatal_required ya valida rol y tipo_paciente

    # ── AUTO-CANCELAR citas pendientes (30 min de margen) ──
    auto_cancelar_citas(Cita.objects.filter(paciente=request.user))
    hoy = timezone.now().date()

    citas = Cita.objects.filter(paciente=request.user).order_by('-fecha')
    citas_pendientes = Cita.objects.filter(
        paciente=request.user, estado__in=['pendiente', 'confirmada'], fecha__gte=hoy
    ).count()

    # Usar template diferente según tipo de paciente
    if request.user.puede_prenatal:
        template = 'paciente/ver_citas_prenatal.html'
    else:
        template = 'paciente/ver_citas_general.html'

    return render(request, template, {
        'citas': citas,
        'user': request.user,
        'citas_pendientes': citas_pendientes,
        'hoy': hoy,
    })
 
@login_required
@no_cache_view
def citas_medico(request):
    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    # ── AUTO-CANCELAR citas pendientes (30 min de margen) ──
    auto_cancelar_citas(Cita.objects.filter(medico=request.user))
    hoy = timezone.now().date()

    citas = Cita.objects.filter(medico=request.user).order_by('fecha', 'hora')

    try:
        es_prenatal = request.user.medico.especialidad and \
                      request.user.medico.especialidad.tipo == 'prenatal'
    except:
        es_prenatal = True

    template = 'medico/citas_medico.html' if es_prenatal else 'medico/citas_medico_general.html'
    return render(request, template, {'citas': citas})
 
@login_required
@no_cache_view
def descargar_pdf_consulta_general(request, consulta_id):
    """Genera y descarga PDF de una consulta general."""
    from paciente_general.models import ConsultaGeneral
    from datetime import datetime
    from django.http import HttpResponse
    
    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    consulta = get_object_or_404(ConsultaGeneral, id=consulta_id)

    # Verificar permisos
    if request.user.rol == 'medico' and consulta.medico_id != request.user.id:
        return redirect('historial_consultas_generales')

    # Usar reportlab para generar PDF
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
        from reportlab.lib import colors
        from io import BytesIO
    except ImportError:
        # Si no tiene reportlab, redirigir a ver_consulta_general
        return redirect('ver_consulta_general', consulta_id=consulta.id)

    # Crear BytesIO
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    styles = getSampleStyleSheet()
    elements = []

    # ── HEADER ──
    header_data = [
        [Paragraph('<b>ZUMEDICAL</b>', ParagraphStyle(name='HeaderTitle', fontSize=16, alignment=1, textColor=colors.HexColor('#b03580'))),
         Paragraph('<b>Centro de Atención Médica</b>', ParagraphStyle(name='HeaderSub', fontSize=9, alignment=1, textColor=colors.HexColor('#8a2563')))],
        [Paragraph('control@zumedical.com | www.zumedical.com', ParagraphStyle(name='HeaderContact', fontSize=8, alignment=1, textColor=colors.grey))]
    ]
    header_table = Table(header_data, colWidths=[3*inch, 3*inch])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.HexColor('#b03580')),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.2*inch))

    # ── TÍTULO ──
    elements.append(Paragraph('<b>REGISTRO DE CONSULTA GENERAL</b>', 
                             ParagraphStyle(name='Title', fontSize=14, alignment=1, textColor=colors.HexColor('#8a2563'), spaceAfter=12)))
    elements.append(Spacer(1, 0.1*inch))

    # ── INFO PACIENTE ──
    info_data = [
        ['<b>Paciente:</b>', f"{consulta.paciente.get_full_name()}", '<b>Fecha:</b>', f"{consulta.fecha.strftime('%d/%m/%Y')}"],
        ['<b>Username:</b>', f"@{consulta.paciente.username}", '<b>Médico:</b>', f"Dr(a). {consulta.medico.get_full_name() if consulta.medico else 'N/A'}"],
    ]
    info_table = Table(info_data, colWidths=[1.2*inch, 1.8*inch, 1.2*inch, 1.8*inch])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e8aad4')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fbf0f7')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.15*inch))

    # ── MOTIVO ──
    if consulta.motivo_consulta:
        elements.append(Paragraph('<b>Motivo de Consulta:</b>', ParagraphStyle(name='SectionTitle', fontSize=10, textColor=colors.HexColor('#8a2563'))))
        elements.append(Paragraph(consulta.motivo_consulta, ParagraphStyle(name='Normal', fontSize=9)))
        elements.append(Spacer(1, 0.1*inch))

    # ── SIGNOS VITALES ──
    if any([consulta.presion_arterial, consulta.temperatura, consulta.peso, consulta.talla]):
        elements.append(Paragraph('<b>Signos Vitales:</b>', ParagraphStyle(name='SectionTitle', fontSize=10, textColor=colors.HexColor('#8a2563'))))
        vitals_data = [
            ['P/A', consulta.presion_arterial or '—', 'Saturación O₂', f"{consulta.saturacion_oxigeno or '—'}%", 'Temp.', f"{consulta.temperatura or '—'}°C"],
            ['FC', f"{consulta.frecuencia_cardiaca or '—'} lpm", 'FR', f"{consulta.frecuencia_respiratoria or '—'} rpm", 'Peso/Talla', f"{consulta.peso or '—'} kg / {consulta.talla or '—'} m"],
        ]
        vitals_table = Table(vitals_data, colWidths=[1*inch, 1.2*inch, 1.2*inch, 1.2*inch, 0.9*inch, 1.5*inch])
        vitals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e8aad4')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fbf0f7')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(vitals_table)
        elements.append(Spacer(1, 0.1*inch))

    # ── DIAGNÓSTICO ──
    if any([consulta.diagnostico_1_patologia, consulta.diagnostico_2_patologia, consulta.diagnostico_3_patologia]):
        elements.append(Paragraph('<b>Diagnóstico:</b>', ParagraphStyle(name='SectionTitle', fontSize=10, textColor=colors.HexColor('#8a2563'))))
        diags_data = [['Patología', 'CIE-10', 'Presuntivo', 'Definitivo']]
        for i in [1, 2, 3]:
            pat = getattr(consulta, f'diagnostico_{i}_patologia')
            if pat:
                diags_data.append([
                    pat,
                    getattr(consulta, f'diagnostico_{i}_cie10') or '—',
                    '✓' if getattr(consulta, f'diagnostico_{i}_presuntivo') else '',
                    '✓' if getattr(consulta, f'diagnostico_{i}_definitivo') else '',
                ])
        diags_table = Table(diags_data, colWidths=[2.5*inch, 1*inch, 0.8*inch, 0.8*inch])
        diags_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e8aad4')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fbf0f7')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(diags_table)
        elements.append(Spacer(1, 0.1*inch))

    # ── PRÓXIMA CITA ──
    if consulta.proxima_cita:
        elements.append(Paragraph('<b>Próxima Cita:</b>', ParagraphStyle(name='SectionTitle', fontSize=10, textColor=colors.HexColor('#8a2563'))))
        proxima_text = consulta.proxima_cita.strftime('%d/%m/%Y')
        if consulta.proxima_cita_hora:
            proxima_text += f" a las {consulta.proxima_cita_hora.strftime('%H:%M')}"
        elements.append(Paragraph(proxima_text, ParagraphStyle(name='Normal', fontSize=9)))
        elements.append(Spacer(1, 0.1*inch))

    # ── FOOTER ──
    elements.append(Spacer(1, 0.2*inch))
    footer_text = f"Documento generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} | Sistema Zumedical"
    elements.append(Paragraph(footer_text, ParagraphStyle(name='Footer', fontSize=7, alignment=1, textColor=colors.grey)))

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    # Return PDF
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f"consulta_{consulta.paciente.username}_{consulta.fecha.strftime('%d-%m-%Y')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)
 
    cita = get_object_or_404(Cita, id=cita_id, medico=request.user)
 
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        if nuevo_estado in ['pendiente', 'confirmada', 'atendido', 'cancelado']:
            cita.estado = nuevo_estado
            cita.save()
            if nuevo_estado in ['cancelado', 'cancelada']:
                registrar_log(request, 'CANCELACION', 'Citas', f"Cambió estado de la cita {cita.id} a {nuevo_estado}", 'WARNING')
            else:
                registrar_log(request, 'UPDATE', 'Citas', f"Cambió estado de la cita {cita.id} a {nuevo_estado}", 'INFO')
            labels = {'confirmada': 'aceptada', 'cancelado': 'rechazada', 'atendido': 'marcada como atendida', 'pendiente': 'vuelta a pendiente'}
            messages.success(request, f'Cita {labels.get(nuevo_estado, "actualizada")} correctamente.')

    return redirect('citas_medico')
 
@login_required
@no_cache_view
def pacientes_medico(request):
    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    from pacientes.models import Paciente
    from citas.models import Cita

    try:
        es_prenatal = request.user.medico.especialidad and \
                      request.user.medico.especialidad.tipo == 'prenatal'
    except Exception:
        es_prenatal = True

    if es_prenatal:
        # ── Médico prenatal: ve sus pacientes asignadas + pacientes con citas
        #    pendientes/atendidas con él que aún no tienen expediente prenatal ──
        #
        # 1) Pacientes que ya tienen medico_prenatal = este médico
        ids_asignadas = Paciente.objects.filter(
            medico_prenatal=request.user
        ).values_list('usuario_id', flat=True)

        # 2) Pacientes con citas con este médico (para poder activar embarazo)
        ids_con_citas = Cita.objects.filter(
            medico=request.user
        ).values_list('paciente_id', flat=True)

        # Unión de ambos conjuntos
        todos_ids = set(ids_asignadas) | set(ids_con_citas)

        pacientes = Paciente.objects.filter(
            usuario_id__in=todos_ids
        ).select_related('usuario', 'medico_prenatal').order_by(
            'usuario__first_name', 'usuario__last_name'
        )

        return render(request, 'medico/pacientes.html', {'pacientes': pacientes})

    else:
        # ── Médico general: ve solo pacientes con citas registradas con él ──
        ids_con_citas = Cita.objects.filter(
            medico=request.user
        ).values_list('paciente_id', flat=True)

        pacientes = Paciente.objects.filter(
            usuario_id__in=ids_con_citas
        ).select_related('usuario').order_by('usuario__first_name', 'usuario__last_name')

        pacientes_activos = pacientes.filter(usuario__is_active=True).count()

        return render(request, 'medico/pacientes_general.html', {
            'pacientes': pacientes,
            'pacientes_activos': pacientes_activos,
            'total': pacientes.count(),
        })

@login_required
def activar_embarazo(request, paciente_id):
    """Activa el embarazo de una paciente desde el panel del médico (método simple)"""
    if request.user.rol not in ['medico', 'enfermera', 'admin']:
        return redirect('landing')

    from pacientes.models import Paciente
    paciente = get_object_or_404(Paciente, id=paciente_id)
    nombre = paciente.usuario.get_full_name()

    if not es_genero_femenino(paciente.usuario.genero):
        messages.error(request, 'Solo las pacientes registradas como femeninas pueden activar embarazo.')
        referer = request.META.get('HTTP_REFERER')
        if referer:
            return redirect(referer)
        return redirect('pacientes_medico')
    
    # Si el género está vacío, asignarlo automáticamente como femenino
    if not paciente.usuario.genero or paciente.usuario.genero.strip() == '':
        paciente.usuario.genero = 'femenino'
        paciente.usuario.save()
    
    # Activar embarazo
    paciente.estado_embarazo = 'ACTIVO'
    paciente.medico_prenatal = request.user
    paciente.mensaje_prenatal_visto = False
    paciente.save()

    registrar_log(request, 'UPDATE', 'Pacientes',
        f'Embarazo activado para {nombre} por {request.user.get_full_name()}', 'INFO')
    messages.success(request, f'Seguimiento prenatal activado para {nombre}.')
    
    # Redirigir al referer o a pacientes
    referer = request.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    return redirect('pacientes_medico')


@login_required
def desactivar_embarazo(request, paciente_id):
    if request.user.rol not in ['medico', 'enfermera', 'admin']:
        return redirect('landing')

    from pacientes.models import Paciente
    paciente = get_object_or_404(Paciente, id=paciente_id)
    nombre = paciente.usuario.get_full_name()
    paciente.estado_embarazo = 'FINALIZADO'
    paciente.mensaje_prenatal_visto = False
    paciente.save()

    registrar_log(request, 'UPDATE', 'Pacientes',
        f'Embarazo desactivado para {nombre} por {request.user.get_full_name()}', 'INFO')
    messages.success(request, f'Módulo prenatal desactivado para {nombre}. La paciente volvió a modo general.')
    referer = request.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    return redirect('pacientes_medico')
@medico_prenatal_required
def registrar_control(request):
    """
    Solo médicos de especialidad PRENATAL pueden registrar controles.
    Al guardar, la IA se ejecuta AUTOMÁTICAMENTE y guarda la predicción.
    """
    try:
        if request.method == 'POST':
            data = request.POST.copy()
            for campo in ('peso', 'altura', 'glucosa', 'temperatura'):
                if campo in data:
                    data[campo] = data.get(campo, '').replace(',', '.')
            form = ControlPrenatalForm(data, medico=request.user)
            if form.is_valid():
                control = form.save(commit=False)
                control.medico = request.user
                control.save()
                registrar_log(request, 'CREATE', 'Controles Prenatales', f"Se registró control para paciente ID {control.paciente.id}", 'INFO')

                # ── Disparar IA automáticamente ────────────────────────────────
                prediccion = _ejecutar_ia_en_control(control, request)

                if prediccion:
                    messages.success(request, f'Control registrado. La IA determinó riesgo {prediccion.nivel_riesgo} ({prediccion.puntuacion_riesgo}%) — revisa el panel clínico completo.')
                    return redirect('ver_predicciones_paciente', paciente_id=control.paciente.id)
                else:
                    messages.success(request, 'Control prenatal registrado. No se pudo generar la evaluación IA automáticamente.')
                    return redirect('historial_prenatal')
        else:
            form = ControlPrenatalForm(medico=request.user)

        return render(request, 'medico/registrar_control.html', {'form': form})
    
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error en registrar_control: {e}", exc_info=True)
        messages.error(request, f'Error al cargar el formulario: {str(e)}')
        return redirect('medico_dashboard')


def _ejecutar_ia_en_control(control, request=None):
    """
    Ejecuta el modelo ML con los datos del control prenatal y guarda PrediccionIA.
    Devuelve el objeto PrediccionIA creado o None si el modelo no está disponible.

    Unidades: glucosa en mg/dL (se convierte internamente a mmol/L para el modelo).
    """
    try:
        from ai.predict import engine as ml_engine
        from prediccion_ia.models import PrediccionIA
        from pacientes.models import Paciente
        import json
        import logging
        logger = logging.getLogger(__name__)

        if not ml_engine.modelo_disponible:
            logger.warning('[IA] Modelo ML no disponible — usando respaldo clinico por reglas')

        # Obtener perfil del paciente
        try:
            perfil = Paciente.objects.get(usuario=control.paciente)
        except Paciente.DoesNotExist:
            logger.error(f'[IA] No existe perfil Paciente para usuario {control.paciente}')
            return None

        def dato_fuera_de_rango(nombre, valor, minimo, maximo, unidad=''):
            if valor < minimo or valor > maximo:
                logger.error(
                    f'[IA] Dato inválido para {nombre}: {valor} {unidad}. '
                    f'Rango permitido: {minimo}-{maximo} {unidad}'
                )
                return True
            return False

        # Edad del paciente. La IA no debe analizar con una edad inventada.
        try:
            edad = int(perfil.edad)
            if dato_fuera_de_rango('edad', edad, 10, 60, 'años'):
                return None
        except (TypeError, ValueError):
            logger.error(f'[IA] Edad inválida para paciente {perfil}: {perfil.edad}')
            return None

        # Parsear presión arterial. No ejecutar IA con valores irreales.
        try:
            sistolica  = int(control.presion_sistolica)
            diastolica = int(control.presion_diastolica)
            if (
                dato_fuera_de_rango('presión sistólica', sistolica, 70, 250, 'mmHg')
                or dato_fuera_de_rango('presión diastólica', diastolica, 40, 150, 'mmHg')
                or sistolica <= diastolica
            ):
                return None
        except Exception:
            logger.error(f'[IA] Presión arterial inválida: {control.presion_arterial}')
            return None

        # BMI calculado desde peso/altura. No reemplazar ceros por valores normales.
        try:
            altura = float(control.altura)
            peso   = float(control.peso)
            if (
                dato_fuera_de_rango('altura', altura, 1.20, 2.20, 'm')
                or dato_fuera_de_rango('peso', peso, 30.0, 250.0, 'kg')
            ):
                return None
            bmi = round(peso / (altura ** 2), 2)
            if dato_fuera_de_rango('IMC', bmi, 10.0, 80.0):
                return None
        except Exception:
            logger.error('[IA] Peso o altura inválidos para calcular IMC')
            return None

        # Glucosa: mg/dL -> mmol/L para el modelo
        try:
            glucosa_mgdl = float(control.glucosa)
            if dato_fuera_de_rango('glucosa', glucosa_mgdl, 40.0, 400.0, 'mg/dL'):
                return None
            glucosa_mmol = round(glucosa_mgdl / 18.0, 3)
        except Exception:
            logger.error(f'[IA] Glucosa inválida: {control.glucosa}')
            return None

        # Frecuencia cardiaca y temperatura con validación. Sin defaults silenciosos.
        try:
            fc = int(control.frecuencia_cardiaca)
            if dato_fuera_de_rango('frecuencia cardíaca', fc, 40, 180, 'lpm'):
                return None
        except Exception:
            logger.error(f'[IA] Frecuencia cardíaca inválida: {control.frecuencia_cardiaca}')
            return None

        try:
            temp = float(control.temperatura)
            if dato_fuera_de_rango('temperatura', temp, 95.0, 106.0, '°F'):
                return None
        except Exception:
            logger.error(f'[IA] Temperatura inválida: {control.temperatura}')
            return None

        # Vector de datos clinicos para el modelo
        datos_clinicos = {
            'age':                  edad,
            'systolic_bp':          sistolica,
            'diastolic_bp':         diastolica,
            'glucose':              glucosa_mmol,
            'body_temp':            temp,
            'heart_rate':           fc,
            'bmi':                  bmi,
            'prev_complications':   int(bool(control.complicaciones_previas)),
            'diabetes_preexisting': int(bool(control.diabetes_preexistente)),
            'diabetes_gestacional': int(bool(control.diabetes_gestacional)),
        }

        logger.info(f'[IA] Prediccion para paciente {perfil} -- datos: {datos_clinicos}')
        try:
            resultado = ml_engine.predict(datos_clinicos)
        except Exception as e:
            logger.error(f'[IA] Error al ejecutar predict: {e}', exc_info=True)
            # Devolver None para permitir que se registre el control sin IA
            return None
        
        logger.info(f'[IA] Resultado: {resultado.nivel_riesgo} ({resultado.puntuacion}%)')

        resultado_json = json.dumps({
            'factores':        resultado.factores_detectados,
            'complicaciones':  resultado.complicaciones,
            'recomendaciones': resultado.recomendaciones,
            'recomendaciones_grupos': resultado.recomendaciones_grupos,
            'probabilidades':  resultado.probabilidades,
            'explicacion':     resultado.explicacion,
            'alerta_critica':  resultado.alerta_critica,
            'nota_antecedente': resultado.nota_antecedente,
            'motor_usado':     'ml_random_forest' if ml_engine.modelo_disponible else 'respaldo_clinico_reglas',
            'datos_procesados': {
                'glucosa_mmol':  glucosa_mmol,
                'glucosa_mgdl':  round(float(control.glucosa), 1),
                'bmi_calculado': bmi,
                'sistolica':     sistolica,
                'diastolica':    diastolica,
            }
        }, ensure_ascii=False)

        prediccion = PrediccionIA.objects.create(
            paciente               = perfil,
            control                = control,
            edad                   = edad,
            semanas_gestacion      = int(control.semanas_gestacion),
            presion_arterial       = str(control.presion_arterial),
            presion_sistolica      = sistolica,
            presion_diastolica     = diastolica,
            peso                   = float(control.peso),
            altura                 = float(control.altura),
            imc                    = bmi,
            glucosa                = float(control.glucosa),
            frecuencia_cardiaca    = fc,
            temperatura            = temp,
            embarazos_previos      = int(control.embarazos_previos),
            complicaciones_previas = bool(control.complicaciones_previas),
            diabetes_preexistente  = bool(control.diabetes_preexistente),
            diabetes_gestacional   = bool(control.diabetes_gestacional),
            nivel_riesgo           = resultado.nivel_riesgo,
            puntuacion_riesgo      = resultado.puntuacion,
            resultado              = resultado_json,
        )
        return prediccion

    except Exception as e:
        import logging, traceback
        logging.getLogger(__name__).error(
            f'[IA] Error al ejecutar prediccion: {e}\n{traceback.format_exc()}'
        )
        return None

@medico_prenatal_required
@login_required
def historial_prenatal(request):
    # medico_prenatal_required ya validó rol y especialidad
 
    paciente_id = request.GET.get('paciente')

    if paciente_id:
        controles = ControlPrenatal.objects.filter(
            paciente_id=paciente_id
        ).select_related('paciente', 'medico').order_by('-fecha')
    else:
        controles = ControlPrenatal.objects.select_related(
            'paciente', 'medico'
        ).order_by('-fecha')

    # ── Enriquecer cada control con su predicción IA ──────────────────────
    from prediccion_ia.models import PrediccionIA
    from pacientes.models import Paciente
    import json

    # Obtener IDs de usuarios de los controles para una sola consulta
    paciente_ids = list({c.paciente_id for c in controles})
    perfiles_map = {
        p.usuario_id: p
        for p in Paciente.objects.filter(usuario_id__in=paciente_ids)
    }

    # Preferir la predicción asociada al control; para datos antiguos sin vínculo,
    # usar la última predicción de la paciente como fallback.
    predicciones_por_control = {}
    for pred in PrediccionIA.objects.filter(
        control_id__in=[c.id for c in controles],
        control_id__isnull=False,
    ).select_related('control').order_by('control_id', '-fecha'):
        if pred.control_id not in predicciones_por_control:
            predicciones_por_control[pred.control_id] = pred

    predicciones_fallback_paciente = {}
    for pred in PrediccionIA.objects.filter(
        paciente__usuario_id__in=paciente_ids,
        control__isnull=True,
    ).select_related('paciente').order_by('paciente', '-fecha'):
        pid = pred.paciente.usuario_id
        if pid not in predicciones_fallback_paciente:
            predicciones_fallback_paciente[pid] = pred

    # Adjuntar al control su predicción clínica correspondiente
    controles_enriquecidos = []
    for c in controles:
        c.prediccion_ia = predicciones_por_control.get(c.id) or predicciones_fallback_paciente.get(c.paciente_id)
        c.imc_calculado = c.imc
        controles_enriquecidos.append(c)

    pacientes = Usuario.objects.filter(
        rol='paciente', paciente__estado_embarazo='ACTIVO'
    ).distinct().order_by('first_name', 'last_name')
 
    return render(request, 'medico/historial.html', {
        'controles': controles_enriquecidos,
        'pacientes': pacientes,
        'paciente_filtrado_id': int(paciente_id) if paciente_id else None,
    })
 
@login_required
def registrar_paciente(request, tipo='prenatal'):
    rol_lower = request.user.rol.lower() if request.user.rol else ''
    if rol_lower not in ['enfermera', 'secretaria']:
        return redireccionar_por_rol(request.user)

    # Validar tipo
    if tipo not in ('prenatal', 'general'):
        tipo = 'prenatal'

    if request.method == 'POST':
        form = RegistroPacienteForm(request.POST)
        if form.is_valid():
            try:
                # El formulario ya crea el usuario y el perfil de paciente
                user = form.save(commit=True)
                
                from pacientes.models import Paciente
                paciente = Paciente.objects.get(usuario=user)
                _aplicar_tipo_paciente(paciente, tipo, user)
                
                # Registrar en log de auditoría
                registrar_log(request, 'CREATE', 'Pacientes',
                    f'{request.user.rol.title()} registró nueva paciente: {user.get_full_name()} (tipo: {tipo})', 'INFO')
                
                messages.success(request, f'Paciente {user.get_full_name()} registrada correctamente como paciente {tipo}.')
                return redirect(f"{reverse('lista_pacientes_enfermera')}?tab={tipo}")
            except Exception as e:
                messages.error(request, f'Error al registrar paciente: {str(e)}')
        else:
            # Mostrar errores del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, error)
                    else:
                        label = form.fields[field].label if field in form.fields else field
                        messages.error(request, f'{label}: {error}')
    else:
        form = RegistroPacienteForm()

    return render(request, 'enfermera/registrar_paciente_enfermera.html', {'form': form, 'tipo': tipo})


@login_required
@no_cache_view
def lista_pacientes_enfermera(request):
    rol_lower = request.user.rol.lower() if request.user.rol else ''
    if rol_lower not in ['enfermera', 'secretaria']:
        return redireccionar_por_rol(request.user)

    from pacientes.models import Paciente
    # Prenatales = tipo prenatal O general con módulo prenatal activado
    pacientes_prenatales = Paciente.objects.filter(
        estado_embarazo='ACTIVO', usuario__rol='paciente'
    ).select_related('usuario').distinct().order_by('-fecha_registro')
    pacientes_generales = Paciente.objects.exclude(
        estado_embarazo='ACTIVO'
    ).filter(usuario__rol='paciente').select_related('usuario').distinct().order_by('-fecha_registro')

    return render(request, 'enfermera/lista_pacientes_enfermera.html', {
        'pacientes_prenatales': pacientes_prenatales,
        'pacientes_generales':  pacientes_generales,
    })


@login_required
def toggle_modulo_prenatal(request, paciente_id):
    """
    La enfermera activa o desactiva el módulo prenatal para una paciente general.
    Solo POST. Solo accesible por enfermera o admin.
    """
    from pacientes.models import Paciente

    rol_lower = request.user.rol.lower() if request.user.rol else ''
    if rol_lower not in ('enfermera', 'secretaria', 'admin'):
        return redireccionar_por_rol(request.user)

    paciente_usuario = get_object_or_404(Usuario, id=paciente_id, rol='paciente')
    perfil, _ = Paciente.objects.get_or_create(usuario=paciente_usuario)

    if request.method == 'POST':
        activar = request.POST.get('activar') == '1'
        if paciente_usuario.paciente.estado_embarazo == 'ACTIVO':
            messages.error(request, 'Solo las pacientes generales usan activación de módulo prenatal.')
            return redirect(request.POST.get('next', 'lista_pacientes_enfermera'))
        if activar and not es_genero_femenino(paciente_usuario.genero):
            messages.error(request, 'Solo las pacientes registradas como femeninas pueden activar el módulo prenatal.')
            return redirect(request.POST.get('next', 'lista_pacientes_enfermera'))

        perfil.estado_embarazo = 'ACTIVO' if activar else 'NINGUNO'
        perfil.save()

        if activar:
            registrar_log(request, 'UPDATE', 'Pacientes',
                f'Módulo prenatal ACTIVADO para {paciente_usuario.get_full_name()} por {request.user.rol}', 'INFO')
            messages.success(
                request,
                f'Módulo prenatal ACTIVADO para {paciente_usuario.get_full_name() or paciente_usuario.username}. '
                f'Ahora puede acceder a controles prenatales, IA y chatbot.'
            )
        else:
            registrar_log(request, 'UPDATE', 'Pacientes',
                f'Módulo prenatal DESACTIVADO para {paciente_usuario.get_full_name()} por {request.user.rol}', 'INFO')
            messages.success(
                request,
                f'Módulo prenatal desactivado para {paciente_usuario.get_full_name() or paciente_usuario.username}.'
            )

    return redirect(request.POST.get('next', 'lista_pacientes_enfermera'))
@login_required
def crear_historia_clinica(request, paciente_id):
    """Crea o edita la Historia Clínica Obstétrica de una paciente."""
    from control_prenatal.models import HistoriaClinica
    from control_prenatal.forms import HistoriaClinicaForm

    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    paciente_usuario = get_object_or_404(Usuario, id=paciente_id, rol='paciente')

    # Si ya existe, redirigir a editar
    historia_existente = HistoriaClinica.objects.filter(paciente=paciente_usuario).first()
    if historia_existente:
        return redirect('editar_historia_clinica', historia_id=historia_existente.id)

    if request.method == 'POST':
        form = HistoriaClinicaForm(request.POST)
        if form.is_valid():
            historia = form.save(commit=False)
            historia.medico = request.user
            historia.save()
            registrar_log(request, 'CREATE', 'Historia Clínica', f"Historia clínica creada para paciente ID {paciente_usuario.id}", 'INFO')
            messages.success(request, f'Historia clínica de {paciente_usuario.get_full_name()} creada correctamente.')
            return redirect('ver_historia_clinica', paciente_id=paciente_id)
    else:
        form = HistoriaClinicaForm(initial={'paciente': paciente_usuario})
        # Bloquear el campo paciente ya que viene predefinido
        form.fields['paciente'].widget.attrs['disabled'] = True

    return render(request, 'medico/historia_clinica_form.html', {
        'form': form,
        'paciente': paciente_usuario,
        'modo': 'crear',
    })


@login_required
def editar_historia_clinica(request, historia_id):
    """Edita una Historia Clínica Obstétrica existente."""
    from control_prenatal.models import HistoriaClinica
    from control_prenatal.forms import HistoriaClinicaForm

    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    historia = get_object_or_404(HistoriaClinica, id=historia_id)

    if request.method == 'POST':
        form = HistoriaClinicaForm(request.POST, instance=historia)
        if form.is_valid():
            form.save()
            registrar_log(request, 'UPDATE', 'Historia Clínica', f"Historia clínica editada ID {historia.id}", 'INFO')
            messages.success(request, 'Historia clínica actualizada correctamente.')
            return redirect('ver_historia_clinica', paciente_id=historia.paciente_id)
    else:
        form = HistoriaClinicaForm(instance=historia)

    return render(request, 'medico/historia_clinica_form.html', {
        'form': form,
        'paciente': historia.paciente,
        'historia': historia,
        'modo': 'editar',
    })


@login_required
def ver_historia_clinica(request, paciente_id):
    """
    Dashboard completo de la Historia Clínica Obstétrica.
    Muestra: Historia inicial + todos los controles + gráfico de ganancia de peso.
    """
    from control_prenatal.models import HistoriaClinica, ControlPrenatal
    from prediccion_ia.models import PrediccionIA
    from pacientes.models import Paciente
    import json

    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    paciente_usuario = get_object_or_404(Usuario, id=paciente_id, rol='paciente')
    perfil = Paciente.objects.filter(usuario=paciente_usuario).first()
    historia = HistoriaClinica.objects.filter(paciente=paciente_usuario).first()

    # Todos los controles ordenados por fecha
    controles = ControlPrenatal.objects.filter(
        paciente=paciente_usuario
    ).order_by('fecha')

    # Última predicción IA
    ultima_prediccion = None
    if perfil:
        ultima_prediccion = PrediccionIA.objects.filter(paciente=perfil).order_by('-fecha').first()

    # ── Datos para el gráfico de ganancia de peso ──────────────────────────
    # Lista de puntos: [{semana, peso, ganancia_desde_inicio}]
    peso_inicial = (historia.peso_inicial if historia and historia.peso_inicial
                    else (controles.first().peso if controles.exists() else None))

    datos_peso_chart = []
    for c in controles:
        ganancia = round(c.peso - peso_inicial, 1) if peso_inicial else 0
        datos_peso_chart.append({
            'semana': c.semanas_gestacion,
            'peso': c.peso,
            'ganancia': ganancia,
            'fecha': c.fecha.strftime('%d/%m/%Y'),
        })

    # Rangos recomendados de ganancia para las bandas del gráfico
    rango_ganancia = None
    if historia:
        rango_ganancia = historia.ganancia_recomendada

    return render(request, 'medico/historia_clinica.html', {
        'paciente':          paciente_usuario,
        'perfil':            perfil,
        'historia':          historia,
        'controles':         controles,
        'ultima_prediccion': ultima_prediccion,
        'datos_peso_json':   json.dumps(datos_peso_chart),
        'rango_ganancia':    rango_ganancia,
        'peso_inicial':      peso_inicial,
    })


@login_required
def editar_perfil_paciente(request, paciente_id):
    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)
 
    from pacientes.models import Paciente
    paciente = get_object_or_404(Paciente, id=paciente_id)
 
    if request.method == 'POST':
        paciente.cedula = request.POST.get('cedula', '')
        paciente.edad = request.POST.get('edad')
        paciente.direccion = request.POST.get('direccion', '')
        paciente.telefono = request.POST.get('telefono', '')
        paciente.fecha_ultima_menstruacion = request.POST.get('fecha_ultima_menstruacion') or None
        paciente.fecha_probable_parto = request.POST.get('fecha_probable_parto') or None
        paciente.save()
        messages.success(request, f'Datos clínicos de {paciente.usuario.get_full_name()} actualizados correctamente.')
        return redirect('pacientes_medico')
 
    return render(request, 'medico/editar_paciente.html', {'paciente': paciente})
 
#HORARIOS DEFINIDOS
# Lunes a Viernes: 8:30 AM - 5:00 PM
HORAS_DISPONIBLES_SEMANA = [
    time(8,30), time(9,0), time(9,30), time(10,0),
    time(10,30), time(11,0), time(11,30),
    time(12,0), time(12,30), time(13,0), time(13,30),
    time(14,0), time(14,30), time(15,0), time(15,30),
    time(16,0), time(16,30), time(17,0)
]

# Sábados y Domingos: 9:00 AM - 3:00 PM
HORAS_DISPONIBLES_FINDE = [
    time(9,0), time(9,30), time(10,0), time(10,30),
    time(11,0), time(11,30), time(12,0), time(12,30),
    time(13,0), time(13,30), time(14,0), time(14,30), time(15,0)
]
 
# AGENDAR CITA (Admin / Enfermera / Secretaria)
@login_required
def agendar_cita_enfermera(request):
    rol_lower = request.user.rol.lower() if request.user.rol else ''
    if rol_lower not in ['admin', 'enfermera', 'secretaria']:
        return redireccionar_por_rol(request.user)

    form = CitaEnfermeraForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            cita = form.save(commit=False)
            existe = Cita.objects.filter(
                medico=cita.medico,
                fecha=cita.fecha,
                hora=cita.hora
            ).exists()
            if existe:
                form.add_error('hora', 'Esta hora ya está ocupada.')
            else:
                # Asignar especialidad del médico automáticamente
                try:
                    from medicos.models import Medico
                    medico_perfil = Medico.objects.get(usuario=cita.medico)
                    if medico_perfil.especialidad:
                        cita.especialidad = medico_perfil.especialidad
                except Exception:
                    pass
                cita.save()
                messages.success(request, 'Cita agendada correctamente.')
                registrar_log(request, 'CREATE', 'Citas',
                    f'Cita agendada para {cita.paciente.get_full_name()} el {cita.fecha} a las {cita.hora}', 'INFO')
                return redirect('todas_citas' if rol_lower == 'admin' else 'citas_enfermera')

    # Pasar médicos con sus especialidades para el JS del template
    from medicos.models import Medico
    import json
    medicos_qs = Medico.objects.select_related('usuario', 'especialidad').filter(
        usuario__is_active=True,
        usuario__rol='medico',
        especialidad__isnull=False,
    ).exclude(
        usuario__username__iregex=r'^(medico_verif|medico_test|test_medico)'
    )
    medicos_data = [
        {
            'usuario_id': m.usuario.id,
            'especialidad__nombre': m.especialidad.nombre if m.especialidad else None,
        }
        for m in medicos_qs
    ]

    return render(request, 'enfermera/agendar_cita_enfermera.html', {
        'form': form,
        'medicos_especialidades': json.dumps(medicos_data),
        'es_admin': rol_lower == 'admin',
    })
 
 
# OBTENER HORAS DISPONIBLES
@login_required
def obtener_horas_disponibles(request):
    from datetime import datetime
    fecha = request.GET.get('fecha')
    medico = request.GET.get('medico')
 
    if not fecha or not medico:
        return JsonResponse({'horas': []})
    
    # Determinar si es fin de semana
    try:
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
        es_finde = fecha_obj.weekday() in [5, 6]  # 5=sábado, 6=domingo
        horas_base = HORAS_DISPONIBLES_FINDE if es_finde else HORAS_DISPONIBLES_SEMANA
    except:
        horas_base = HORAS_DISPONIBLES_SEMANA
 
    horas_ocupadas = Cita.objects.filter(
        fecha=fecha,
        medico_id=medico,
        estado__in=['pendiente', 'confirmada']
    ).values_list('hora', flat=True)
 
    horas_libres = [
        h.strftime("%H:%M")
        for h in horas_base if h not in horas_ocupadas
    ]
 
    return JsonResponse({'horas': horas_libres})
@login_required
def editar_paciente_enfermera(request, paciente_id):
    rol_lower = request.user.rol.lower() if request.user.rol else ''
    if rol_lower not in ['enfermera', 'secretaria']:
        return redireccionar_por_rol(request.user)
    paciente = get_object_or_404(Usuario, id=paciente_id, rol='paciente')
    from pacientes.models import Paciente
    perfil, _ = Paciente.objects.get_or_create(usuario=paciente)
    if request.method == 'POST':
        form = EditarPacienteEnfermeraForm(request.POST, instance=perfil, usuario=paciente)
        if form.is_valid():
            # Validar cédula duplicada (excluir el propio paciente)
            cedula_nueva = form.cleaned_data.get('cedula', '').strip()
            if cedula_nueva:
                duplicado = Paciente.objects.filter(cedula=cedula_nueva).exclude(usuario=paciente).first()
                if duplicado:
                    messages.error(request, f"¡La cédula {cedula_nueva} ya está registrada para otro paciente ({duplicado.usuario.get_full_name()}).")
                    return render(request, 'enfermera/editar_paciente_enfermera.html', {
                        'form': form, 'paciente': paciente, 'error_cedula': True
                    })
            username_nuevo = form.cleaned_data.get('username', '').strip()
            if Usuario.objects.filter(username__iexact=username_nuevo).exclude(id=paciente.id).exists():
                messages.error(request, f'El usuario "{username_nuevo}" ya está registrado.')
                return render(request, 'enfermera/editar_paciente_enfermera.html', {
                    'form': form, 'paciente': paciente
                })
            form.save(usuario=paciente)
            # Actualizar cédula y teléfono en modelo Paciente
            if cedula_nueva:
                perfil.cedula = cedula_nueva
            telefono_nuevo = form.cleaned_data.get('telefono', '').strip()
            if telefono_nuevo:
                perfil.telefono = telefono_nuevo
            perfil.save()
            registrar_log(request, 'UPDATE', 'Secretaría', f"Se editó el paciente {paciente.get_full_name()} (ID: {paciente.id})", 'INFO')
            messages.success(request, f"Paciente {paciente.get_full_name()} actualizado correctamente.")
            return redirect('lista_pacientes_enfermera')
        else:
            messages.error(request, "Por favor corrige los errores del formulario.")
    else:
        form = EditarPacienteEnfermeraForm(instance=perfil, usuario=paciente)
    
    return render(request, 'enfermera/editar_paciente_enfermera.html', {
        'form': form,
        'paciente': paciente
    })

@login_required
def reprogramar_cita(request, cita_id):
    rol_lower = request.user.rol.lower() if request.user.rol else ''
    if rol_lower not in ['enfermera', 'secretaria']:
        return redireccionar_por_rol(request.user)
 
    try:
        cita = Cita.objects.get(id=cita_id)
    except Cita.DoesNotExist:
        messages.error(request, 'La cita no existe.')
        return redirect('citas_enfermera')
 
    if request.method == 'POST':
        nueva_fecha = request.POST.get('fecha')
        nueva_hora = request.POST.get('hora')
        
        # Validar que los campos no estén vacíos
        if not nueva_fecha or not nueva_hora:
            messages.error(request, 'Por favor completa la fecha y la hora.')
            return render(request, 'enfermera/reprogramar_enfermera.html', {
                'cita': cita,
                'hoy': timezone.now().date(),
            })
        
        # Validar que no exista otra cita en esa fecha y hora para el mismo médico
        cita_existente = Cita.objects.filter(
            medico=cita.medico,
            fecha=nueva_fecha,
            hora=nueva_hora
        ).exclude(id=cita_id).exists()
        
        if cita_existente:
            messages.error(request, 'Ya existe una cita para este médico en esa fecha y hora. Por favor selecciona otro horario.')
            return render(request, 'enfermera/reprogramar_enfermera.html', {
                'cita': cita,
                'hoy': timezone.now().date(),
            })
        
        try:
            from datetime import datetime
            # Convertir strings a tipos correctos
            cita.fecha = datetime.strptime(nueva_fecha, '%Y-%m-%d').date()
            cita.hora = datetime.strptime(nueva_hora, '%H:%M').time()
            cita.save()
            registrar_log(request, 'UPDATE', 'Citas', f"Enfermera reprogramó cita {cita.id} a {cita.fecha} {cita.hora}", 'INFO')
            messages.success(request, 'Cita reprogramada correctamente.')
            return redirect('citas_enfermera')
        except Exception as e:
            messages.error(request, f'Error al reprogramar la cita: {str(e)}')
            return render(request, 'enfermera/reprogramar_enfermera.html', {
                'cita': cita,
                'hoy': timezone.now().date(),
            })
 
    return render(request, 'enfermera/reprogramar_enfermera.html', {
        'cita': cita,
        'hoy': timezone.now().date(),
    })

@login_required
@login_required
@no_cache_view
def cancelar_cita_enfermera(request, cita_id):
    """Permite a la enfermera cancelar una cita con motivo."""
    rol_lower = request.user.rol.lower() if request.user.rol else ''
    if rol_lower not in ['enfermera', 'secretaria']:
        return redireccionar_por_rol(request.user)
    
    cita = get_object_or_404(Cita, id=cita_id)
    
    if request.method == 'POST':
        motivo_cancelacion = request.POST.get('motivo_cancelacion', '').strip()
        cita.estado = 'cancelada'
        cita.motivo_cancelacion = motivo_cancelacion
        cita.save()
        registrar_log(request, 'CANCELACION', 'Citas',
            f'Cita #{cita.id} cancelada por enfermera. Paciente: {cita.paciente.get_full_name()}. Motivo: {motivo_cancelacion}', 'INFO')
        messages.success(request, 'Cita cancelada correctamente.')
        return redirect('citas_enfermera')
    
    return render(request, 'enfermera/cancelar_cita_modal.html', {'cita': cita})
 
@login_required
@no_cache_view
def citas_enfermera(request):
    rol_lower = request.user.rol.lower() if request.user.rol else ''
    if rol_lower not in ['enfermera', 'secretaria']:
        return redireccionar_por_rol(request.user)

    # ── AUTO-CANCELAR citas pendientes (30 min de margen) ──
    auto_cancelar_citas(Cita.objects)
    hoy = timezone.now().date()

    citas_prenatales = Cita.objects.filter(
        medico__medico__especialidad__tipo='prenatal'
    ).select_related('paciente', 'medico', 'especialidad').distinct().order_by('-fecha', 'hora')

    citas_generales = Cita.objects.exclude(
        medico__medico__especialidad__tipo='prenatal'
    ).select_related('paciente', 'medico', 'especialidad').order_by('-fecha', 'hora')

    return render(request, 'enfermera/citas_enfermera.html', {
        'citas_prenatales': citas_prenatales,
        'citas_generales': citas_generales,
        'total_prenatales': citas_prenatales.count(),
        'total_generales': citas_generales.count(),
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })
 
@login_required
def datos_paciente(request):
    paciente_id = request.GET.get('paciente_id')
    if not paciente_id:
        return JsonResponse({'error': 'No se proporcionó paciente'}, status=400)
 
    try:
        from pacientes.models import Paciente
        paciente = Paciente.objects.get(usuario_id=paciente_id)
        return JsonResponse({
            'cedula': paciente.cedula or '',
            'telefono': paciente.telefono or '',
            'email': paciente.usuario.email or '',
        })
    except Paciente.DoesNotExist:
        return JsonResponse({'cedula': '', 'telefono': '', 'email': ''})
 
@login_required
def buscar_pacientes(request):
    """
    Endpoint AJAX para el buscador en tiempo real de pacientes.
    Busca por nombre, apellido o cédula.
    Retorna: {"pacientes": [{"id": ..., "nombre": ..., "cedula": ..., "email": ...}]}
    """
    if request.user.rol not in ('enfermera', 'admin'):
        return JsonResponse({'pacientes': []}, status=403)
 
    q = request.GET.get('q', '').strip()
    if not q or len(q) < 1:
        return JsonResponse({'pacientes': []})
 
    from pacientes.models import Paciente
    from django.db.models import Q
 
    pacientes = Paciente.objects.filter(
        Q(usuario__first_name__icontains=q) |
        Q(usuario__last_name__icontains=q)  |
        Q(cedula__icontains=q)              |
        Q(usuario__email__icontains=q)
    ).select_related('usuario')[:15]
 
    resultados = []
    for p in pacientes:
        nombre_completo = f"{p.usuario.first_name} {p.usuario.last_name}".strip()
        if not nombre_completo:
            nombre_completo = p.usuario.username
        resultados.append({
            'id':     p.usuario.id,   # mismo ID que usa datos_paciente (usuario_id)
            'nombre': nombre_completo,
            'cedula': p.cedula or '',
            'email':  p.usuario.email or '',
        })
 
    return JsonResponse({'pacientes': resultados})
 
#ADMIN
@login_required
@no_cache_view
def lista_usuarios(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)
 
    usuarios = User.objects.all()
 
    # Obtener IDs de usuarios con sesión activa
    sesiones_activas = Session.objects.filter(expire_date__gte=timezone.now())
    ids_activos = set()
    for sesion in sesiones_activas:
        datos = sesion.get_decoded()
        uid = datos.get('_auth_user_id')
        if uid:
            ids_activos.add(int(uid))
 
    # Agregar flag a cada usuario
    for u in usuarios:
        u.sesion_activa = u.id in ids_activos
 
    return render(request, 'admin/lista_usuarios.html', {
        'usuarios': usuarios,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })
 
 
# ─────────────────────────────────────────────────────────────
# AGREGAR ESTAS DOS VIEWS NUEVAS
# ─────────────────────────────────────────────────────────────
 
@login_required
def toggle_usuario(request, usuario_id):
    """Activa o desactiva la cuenta de un usuario (solo admin, no se puede tocar a otro admin)."""
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)
 
    if request.method != 'POST':
        return redirect('lista_usuarios')
 
    usuario = get_object_or_404(User, id=usuario_id)
 
    # Protección: no se puede desactivar a otro admin ni a uno mismo
    if usuario.rol == 'admin':
        messages.error(request, 'No se puede desactivar la cuenta de un administrador.')
        return redirect('lista_usuarios')
 
    if usuario == request.user:
        messages.error(request, 'No puedes desactivar tu propia cuenta.')
        return redirect('lista_usuarios')
 
    accion = request.POST.get('accion')
 
    if accion == 'desactivar':
        usuario.is_active = False
        usuario.save()
        # Cerrar todas las sesiones activas del usuario
        sesiones = Session.objects.filter(expire_date__gte=timezone.now())
        for sesion in sesiones:
            datos = sesion.get_decoded()
            if datos.get('_auth_user_id') == str(usuario.id):
                sesion.delete()
        messages.success(request, f'La cuenta de {usuario.get_full_name() or usuario.username} ha sido desactivada.')
        registrar_log(request, 'UPDATE', 'Usuarios',
            f'Cuenta de "{usuario.username}" desactivada — sesiones cerradas forzosamente', 'WARNING')

    elif accion == 'activar':
        usuario.is_active = True
        usuario.save()
        messages.success(request, f'La cuenta de {usuario.get_full_name() or usuario.username} ha sido reactivada.')
        registrar_log(request, 'UPDATE', 'Usuarios',
            f'Cuenta de "{usuario.username}" (ID {usuario.id}) reactivada', 'INFO')
 
    return redirect('lista_usuarios')
 
 
@login_required
@no_cache_view
def lista_medicos(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)
 
    from medicos.models import Medico
    medicos = Medico.objects.all()
    return render(request, 'admin/lista_medicos.html', {
        'medicos': medicos,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })
 
 
@login_required
@no_cache_view
def lista_pacientes(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    from pacientes.models import Paciente
    pacientes_prenatales = Paciente.objects.filter(
        estado_embarazo='ACTIVO', usuario__rol='paciente'
    ).select_related('usuario').distinct()
    pacientes_generales = Paciente.objects.exclude(
        estado_embarazo='ACTIVO'
    ).filter(usuario__rol='paciente').select_related('usuario').distinct()

    return render(request, 'admin/lista_pacientes.html', {
        'pacientes_prenatales': pacientes_prenatales,
        'pacientes_generales': pacientes_generales,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })
 
 
@login_required
@no_cache_view
def todas_citas(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    try:
        # ── AUTO-CANCELAR citas pendientes (30 min de margen) ──
        try:
            auto_cancelar_citas(Cita.objects.all())
        except Exception as e:
            print(f"Error al auto-cancelar citas: {e}")

        # Obtener citas prenatales
        citas_prenatales = Cita.objects.filter(
            medico__medico__especialidad__tipo='prenatal'
        ).select_related('paciente', 'medico', 'especialidad').distinct().order_by('-fecha', 'hora')
        
        print(f"DEBUG: Citas prenatales encontradas: {citas_prenatales.count()}")

        # Obtener citas generales
        citas_generales = Cita.objects.exclude(
            medico__medico__especialidad__tipo='prenatal'
        ).select_related('paciente', 'medico', 'especialidad').order_by('-fecha', 'hora')
        
        print(f"DEBUG: Citas generales encontradas: {citas_generales.count()}")

        return render(request, 'admin/todas_citas.html', {
            'citas_prenatales': citas_prenatales,
            'citas_generales': citas_generales,
            'total_prenatales': citas_prenatales.count(),
            'total_generales': citas_generales.count(),
            'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
        })
    except Exception as e:
        print(f"ERROR COMPLETO EN todas_citas: {e}")
        import traceback
        traceback.print_exc()
        raise
 
 
@login_required
@no_cache_view
def controles_admin(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)
 
    controles = ControlPrenatal.objects.all()
    return render(request, 'admin/controles.html', {
        'controles': controles,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


DOMINIOS_EMAIL_PERMITIDOS = {
    'gmail.com', 'googlemail.com', 'outlook.com', 'hotmail.com', 'live.com',
    'msn.com', 'yahoo.com', 'yahoo.es', 'icloud.com', 'me.com', 'mac.com',
    'proton.me', 'protonmail.com', 'aol.com', 'zoho.com', 'gmx.com',
    'gmx.net', 'mail.com', 'yandex.com', 'yandex.ru', 'fastmail.com',
    'tutanota.com', 'tuta.com', 'hey.com', 'inbox.com',
}


def _admin_citas_pendientes_count():
    try:
        return Cita.objects.filter(estado='pendiente').count()
    except Exception:
        logger.exception("Error al contar citas pendientes para el panel admin")
        return 0


def _validar_usuario_admin(first_name, last_name, username, email, password, rol=None, roles_validos=None, password_confirm=None):
    import re
    from django.core.exceptions import ValidationError
    from django.core.validators import validate_email

    errores = []
    nombre_re = re.compile(r"^[A-Za-zÁÉÍÓÚÜÑáéíóúüñ\s]{2,40}$")
    usuario_re = re.compile(r"^[A-Za-z0-9._]{3,30}$")

    if not nombre_re.fullmatch(first_name or ""):
        errores.append("El nombre debe tener de 2 a 40 caracteres y solo letras.")
    if not nombre_re.fullmatch(last_name or ""):
        errores.append("Los apellidos deben tener de 2 a 40 caracteres y solo letras.")
    if not usuario_re.fullmatch(username or ""):
        errores.append("El usuario debe tener de 3 a 30 caracteres, sin espacios ni símbolos especiales.")
    if email:
        try:
            validate_email(email)
        except ValidationError:
            errores.append("Ingresa un correo electrónico válido.")
        else:
            dominio = email.rsplit('@', 1)[-1].lower()
            if dominio not in DOMINIOS_EMAIL_PERMITIDOS:
                errores.append("El correo debe usar un dominio permitido como gmail.com, outlook.com, hotmail.com o yahoo.com.")
    if roles_validos is not None and rol not in roles_validos:
        errores.append("Selecciona un rol válido.")
    if len(password or "") < 8 or not re.search(r"[A-ZÁÉÍÓÚÑ]", password or "") or not re.search(r"\d", password or "") or not re.search(r"[^A-Za-z0-9ÁÉÍÓÚÑáéíóúñ]", password or ""):
        errores.append("La contraseña debe tener mínimo 8 caracteres, una mayúscula, un número y un símbolo.")
    if password_confirm is not None and password != password_confirm:
        errores.append("La confirmación de contraseña no coincide.")

    return errores


def _tabla_tiene_columna(tabla, columna):
    try:
        with connection.cursor() as cursor:
            columnas = connection.introspection.get_table_description(cursor, tabla)
        return any(col.name == columna for col in columnas)
    except Exception:
        logger.exception("No se pudo inspeccionar la tabla %s", tabla)
        return False


def _crear_perfil_medico_compatible(usuario, especialidad_obj, telefono):
    from medicos.models import Medico

    especialidad_nombre = getattr(especialidad_obj, 'nombre', '') or ''
    tiene_columna_legacy = _tabla_tiene_columna('medicos_medico', 'especialidad')
    tiene_columna_fk = _tabla_tiene_columna('medicos_medico', 'especialidad_id')

    if tiene_columna_legacy and tiene_columna_fk:
        with connection.cursor() as cursor:
            cursor.execute(
                "INSERT INTO medicos_medico (usuario_id, especialidad_id, telefono, especialidad) VALUES (%s, %s, %s, %s)",
                [usuario.id, especialidad_obj.id if especialidad_obj else None, telefono, especialidad_nombre[:100]],
            )
        return Medico.objects.get(usuario=usuario)

    return Medico.objects.create(
        usuario=usuario,
        especialidad=especialidad_obj,
        telefono=telefono,
    )


def _validar_entero_opcional(valor, etiqueta, minimo=1, maximo=120):
    if valor in (None, ''):
        return None, None
    valor = str(valor).strip()
    if not valor.isdigit():
        return None, f'{etiqueta} debe ser un número válido.'
    numero = int(valor)
    if numero < minimo or numero > maximo:
        return None, f'{etiqueta} debe estar entre {minimo} y {maximo}.'
    return numero, None


def _aplicar_tipo_paciente(paciente, tipo, usuario=None):
    """Mantiene consistente la clasificación usada por los listados."""
    if tipo == 'prenatal':
        paciente.estado_embarazo = 'ACTIVO'
        paciente.mensaje_prenatal_visto = False
        if usuario and usuario.genero != 'femenino':
            usuario.genero = 'femenino'
            usuario.save(update_fields=['genero'])
    else:
        paciente.estado_embarazo = 'NINGUNO'
        paciente.medico_prenatal = None
        paciente.mensaje_prenatal_visto = False
    paciente.save()




@login_required
def admin_crear_usuario(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    from landing.models import Especialidad

    try:
        especialidades = Especialidad.objects.filter(activo=True)
        list(especialidades[:1])
    except Exception:
        logger.exception("Error al cargar especialidades en crear usuario")
        especialidades = Especialidad.objects.none()
        messages.error(request, 'No se pudieron cargar las especialidades médicas. Revisa la base de datos o las migraciones en producción.')
    roles = ['admin', 'medico', 'enfermera', 'paciente']

    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        rol        = request.POST.get('rol', '')
        password   = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        errores = _validar_usuario_admin(first_name, last_name, username, email, password, rol, roles, password_confirm)
        especialidad_obj = None

        if rol == 'medico':
            especialidad_id = request.POST.get('especialidad_id', '').strip()
            telefono_med = request.POST.get('telefono_med', '').strip()
            if not especialidad_id:
                errores.append("La especialidad del médico es obligatoria.")
            else:
                try:
                    especialidad_obj = especialidades.get(id=especialidad_id)
                except Especialidad.DoesNotExist:
                    errores.append("Selecciona una especialidad médica válida.")
            if telefono_med and (not telefono_med.isdigit() or len(telefono_med) != 10):
                errores.append("El teléfono del médico debe tener 10 dígitos.")

        if errores:
            for error in errores:
                messages.error(request, error)
            return render(request, 'admin/crear_usuario.html', {
                'roles': roles,
                'especialidades': especialidades,
                'form_data': request.POST,
                'citas_pendientes': _admin_citas_pendientes_count(),
            })

        if User.objects.filter(username=username).exists():
            messages.error(request, f'El usuario "{username}" ya existe.')
            return render(request, 'admin/crear_usuario.html', {
                'roles': roles,
                'especialidades': especialidades,
                'form_data': request.POST,
                'citas_pendientes': _admin_citas_pendientes_count(),
            })
        if email and User.objects.filter(email__iexact=email).exists():
            messages.error(request, f'El correo "{email}" ya está registrado.')
            return render(request, 'admin/crear_usuario.html', {
                'roles': roles,
                'especialidades': especialidades,
                'form_data': request.POST,
                'citas_pendientes': _admin_citas_pendientes_count(),
            })

        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    username=username, password=password,
                    first_name=first_name, last_name=last_name,
                    email=email, rol=rol
                )

                # Crear perfiles según rol
                if rol == 'medico':
                    _crear_perfil_medico_compatible(
                        user,
                        especialidad_obj,
                        request.POST.get('telefono_med', '').strip()
                    )
                elif rol == 'paciente':
                    from pacientes.models import Paciente
                    Paciente.objects.get_or_create(usuario=user)
        except Exception:
            logger.exception("Error al crear usuario desde el panel admin")
            messages.error(request, 'No se pudo crear el usuario. Revisa los datos e intenta nuevamente.')
            return render(request, 'admin/crear_usuario.html', {
                'roles': roles,
                'especialidades': especialidades,
                'form_data': request.POST,
                'citas_pendientes': _admin_citas_pendientes_count(),
            })

        messages.success(request, f'Usuario "{username}" creado correctamente.')
        registrar_log(request, 'CREATE', 'Usuarios',
            f'Usuario "{username}" creado con rol "{rol}"', 'INFO')
        return redirect('lista_usuarios')

    return render(request, 'admin/crear_usuario.html', {
        'roles': roles,
        'especialidades': especialidades,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


@login_required
def admin_editar_usuario(request, usuario_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    usuario = get_object_or_404(User, id=usuario_id)
    from landing.models import Especialidad

    especialidades = Especialidad.objects.filter(activo=True)

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        if email and User.objects.filter(email__iexact=email).exclude(id=usuario.id).exists():
            messages.error(request, f'El correo "{email}" ya está registrado.')
            return render(request, 'admin/editar_usuario.html', {
                'usuario': usuario,
                'roles': ['admin', 'medico', 'enfermera', 'paciente'],
                'especialidades': especialidades,
                'medico_perfil': getattr(usuario, 'medico', None) if usuario.rol == 'medico' else None,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })
        nuevo_rol = request.POST.get('rol', usuario.rol)
        telefono_med = request.POST.get('telefono_med', '').strip()
        especialidad_obj = None

        if nuevo_rol == 'medico':
            especialidad_id = request.POST.get('especialidad_id', '').strip()
            if not especialidad_id:
                messages.error(request, 'La especialidad del médico es obligatoria.')
                return render(request, 'admin/editar_usuario.html', {
                    'usuario': usuario,
                    'roles': ['admin', 'medico', 'enfermera', 'paciente'],
                    'especialidades': especialidades,
                    'medico_perfil': getattr(usuario, 'medico', None),
                    'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
                })
            try:
                especialidad_obj = especialidades.get(id=especialidad_id)
            except Especialidad.DoesNotExist:
                messages.error(request, 'Selecciona una especialidad médica válida.')
                return render(request, 'admin/editar_usuario.html', {
                    'usuario': usuario,
                    'roles': ['admin', 'medico', 'enfermera', 'paciente'],
                    'especialidades': especialidades,
                    'medico_perfil': getattr(usuario, 'medico', None),
                    'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
                })
            if telefono_med and (not telefono_med.isdigit() or len(telefono_med) != 10):
                messages.error(request, 'El teléfono del médico debe tener 10 dígitos.')
                return render(request, 'admin/editar_usuario.html', {
                    'usuario': usuario,
                    'roles': ['admin', 'medico', 'enfermera', 'paciente'],
                    'especialidades': especialidades,
                    'medico_perfil': getattr(usuario, 'medico', None),
                    'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
                })

        with transaction.atomic():
            usuario.first_name = request.POST.get('first_name', '').strip()
            usuario.last_name  = request.POST.get('last_name', '').strip()
            usuario.email      = email
            usuario.rol = nuevo_rol
            password = request.POST.get('password', '').strip()
            if password:
                usuario.set_password(password)
            usuario.save()

            # Actualizar perfil médico si aplica
            if nuevo_rol == 'medico':
                from medicos.models import Medico
                medico, _ = Medico.objects.get_or_create(usuario=usuario)
                medico.especialidad = especialidad_obj
                medico.telefono     = telefono_med
                medico.save()

        messages.success(request, f'Usuario "{usuario.username}" actualizado correctamente.')
        registrar_log(request, 'UPDATE', 'Usuarios',
            f'Usuario "{usuario.username}" (ID {usuario.id}) actualizado — rol: {nuevo_rol}', 'INFO')
        return redirect('lista_usuarios')

    medico_perfil = None
    if usuario.rol == 'medico':
        from medicos.models import Medico
        try:
            medico_perfil = Medico.objects.get(usuario=usuario)
        except Medico.DoesNotExist:
            pass

    return render(request, 'admin/editar_usuario.html', {
        'usuario': usuario,
        'roles': ['admin', 'medico', 'enfermera', 'paciente'],
        'especialidades': especialidades,
        'medico_perfil': medico_perfil,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


@login_required
def admin_eliminar_usuario(request, usuario_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    usuario = get_object_or_404(User, id=usuario_id)

    if usuario == request.user:
        messages.error(request, 'No puedes eliminar tu propia cuenta.')
        return redirect('lista_usuarios')

    if request.method == 'POST':
        nombre = usuario.username
        usuario.delete()
        messages.success(request, f'Usuario "{nombre}" eliminado correctamente.')
        registrar_log(request, 'DELETE', 'Usuarios',
            f'Usuario "{nombre}" (ID {usuario_id}) eliminado del sistema', 'WARNING')
        return redirect('lista_usuarios')

    return render(request, 'admin/confirmar_eliminar.html', {
        'objeto': usuario.get_full_name() or usuario.username,
        'tipo': 'usuario',
        'volver': 'lista_usuarios',
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


# ── PACIENTES ─────────────────────────────────

@login_required
def admin_crear_paciente(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        password   = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        genero     = 'femenino'
        cedula     = request.POST.get('cedula', '').strip()
        telefono   = request.POST.get('telefono', '').strip()
        edad, edad_error = _validar_entero_opcional(request.POST.get('edad'), 'La edad', 1, 99)

        errores = _validar_usuario_admin(first_name, last_name, username, email, password, password_confirm=password_confirm)
        if cedula and (not cedula.isdigit() or len(cedula) != 10):
            errores.append('La cédula debe tener exactamente 10 números.')
        if telefono and (not telefono.isdigit() or len(telefono) != 10):
            errores.append('El teléfono debe tener exactamente 10 números.')
        if edad_error:
            errores.append(edad_error)
        if errores:
            for error in errores:
                messages.error(request, error)
            return render(request, 'admin/crear_paciente.html', {
                'form_data': request.POST,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })

        from pacientes.models import Paciente

        if User.objects.filter(username=username).exists():
            messages.error(request, f'El usuario "{username}" ya existe.')
            return render(request, 'admin/crear_paciente.html', {
                'form_data': request.POST,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })
        if email and User.objects.filter(email__iexact=email).exists():
            messages.error(request, f'El correo "{email}" ya está registrado.')
            return render(request, 'admin/crear_paciente.html', {
                'form_data': request.POST,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })
        if cedula and Paciente.objects.filter(cedula=cedula).exists():
            messages.error(request, f'La cédula "{cedula}" ya está registrada.')
            return render(request, 'admin/crear_paciente.html', {
                'form_data': request.POST,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })

        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    username=username, password=password,
                    first_name=first_name, last_name=last_name,
                    email=email, rol='paciente', genero=genero
                )
                paciente, _ = Paciente.objects.get_or_create(usuario=user)
                paciente.cedula    = cedula
                paciente.telefono  = telefono
                paciente.edad      = edad
                paciente.direccion = request.POST.get('direccion', '')
                fecha_um = request.POST.get('fecha_ultima_menstruacion')
                fecha_pp = request.POST.get('fecha_probable_parto')
                paciente.fecha_ultima_menstruacion = fecha_um if fecha_um else None
                paciente.fecha_probable_parto      = fecha_pp if fecha_pp else None
                _aplicar_tipo_paciente(paciente, 'prenatal', user)
        except Exception:
            logger.exception("Error al crear paciente prenatal desde el panel admin")
            messages.error(request, 'No se pudo crear la paciente. Revisa los datos e intenta nuevamente.')
            return render(request, 'admin/crear_paciente.html', {
                'form_data': request.POST,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })

        messages.success(request, f'Paciente "{first_name} {last_name}" creada correctamente.')
        registrar_log(request, 'CREATE', 'Pacientes',
            f'Paciente "{first_name} {last_name}" (usuario: {username}) registrada', 'INFO')
        return redirect(f"{reverse('lista_pacientes')}?tab=prenatal")

    return render(request, 'admin/crear_paciente.html', {
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


@login_required
def admin_editar_paciente(request, paciente_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    from pacientes.models import Paciente
    paciente = get_object_or_404(Paciente, id=paciente_id)

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        if username and Usuario.objects.filter(username__iexact=username).exclude(id=paciente.usuario.id).exists():
            messages.error(request, f'El usuario "{username}" ya está registrado.')
            return render(request, 'admin/editar_paciente.html', {
                'paciente': paciente,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })
        if username:
            paciente.usuario.username = username
        email = request.POST.get('email', '').strip()
        if email and Usuario.objects.filter(email__iexact=email).exclude(id=paciente.usuario.id).exists():
            messages.error(request, f'El correo "{email}" ya está registrado.')
            return render(request, 'admin/editar_paciente.html', {
                'paciente': paciente,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })
        paciente.usuario.first_name = request.POST.get('first_name', '').strip()
        paciente.usuario.last_name  = request.POST.get('last_name', '').strip()
        paciente.usuario.email      = email
        password = request.POST.get('password', '').strip()
        if password:
            paciente.usuario.set_password(password)
        paciente.usuario.save()

        paciente.cedula    = request.POST.get('cedula', '')
        paciente.telefono  = request.POST.get('telefono', '')
        paciente.edad      = request.POST.get('edad') or None
        paciente.direccion = request.POST.get('direccion', '')
        fecha_um = request.POST.get('fecha_ultima_menstruacion')
        fecha_pp = request.POST.get('fecha_probable_parto')
        paciente.fecha_ultima_menstruacion = fecha_um if fecha_um else None
        paciente.fecha_probable_parto      = fecha_pp if fecha_pp else None
        paciente.save()

        messages.success(request, 'Paciente actualizada correctamente.')
        registrar_log(request, 'UPDATE', 'Pacientes',
            f'Datos de paciente "{paciente.usuario.get_full_name()}" (ID {paciente.id}) actualizados', 'INFO')
        tab = 'prenatal' if paciente.estado_embarazo == 'ACTIVO' else 'general'
        return redirect(f"{reverse('lista_pacientes')}?tab={tab}")

    return render(request, 'admin/editar_paciente.html', {
        'paciente': paciente,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


@login_required
def admin_eliminar_paciente(request, paciente_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    from pacientes.models import Paciente
    paciente = get_object_or_404(Paciente, id=paciente_id)

    if request.method == 'POST':
        nombre = str(paciente)
        tab = 'prenatal' if paciente.estado_embarazo == 'ACTIVO' else 'general'
        paciente.usuario.delete()
        messages.success(request, f'Paciente "{nombre}" eliminada correctamente.')
        registrar_log(request, 'DELETE', 'Pacientes',
            f'Paciente "{nombre}" (ID {paciente_id}) eliminada del sistema', 'WARNING')
        return redirect(f"{reverse('lista_pacientes')}?tab={tab}")

    return render(request, 'admin/confirmar_eliminar.html', {
        'objeto': f'{paciente.usuario.first_name} {paciente.usuario.last_name}',
        'tipo': 'paciente',
        'volver': 'lista_pacientes',
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


# ── MÉDICOS ───────────────────────────────────
@login_required
def admin_crear_medico(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    from medicos.models import Medico
    from landing.models import Especialidad

    try:
        especialidades = Especialidad.objects.filter(activo=True)
        list(especialidades[:1])
    except Exception:
        logger.exception("Error al cargar especialidades en crear médico")
        especialidades = Especialidad.objects.none()
        messages.error(request, 'No se pudieron cargar las especialidades médicas. Revisa la base de datos o las migraciones en producción.')

    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        password   = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        telefono   = request.POST.get('telefono', '').strip()
        especialidad_id = request.POST.get('especialidad_id', '').strip()
        errores = _validar_usuario_admin(first_name, last_name, username, email, password, password_confirm=password_confirm)
        especialidad_obj = None

        if telefono and (not telefono.isdigit() or len(telefono) != 10):
            errores.append("El teléfono debe tener 10 dígitos.")
        if not especialidad_id:
            errores.append("Selecciona una especialidad médica.")
        else:
            try:
                especialidad_obj = especialidades.get(id=especialidad_id)
            except Especialidad.DoesNotExist:
                errores.append("Selecciona una especialidad médica válida.")

        if errores:
            for error in errores:
                messages.error(request, error)
            return render(request, 'admin/crear_medico.html', {
                'form_data': request.POST,
                'especialidades': especialidades,
                'citas_pendientes': _admin_citas_pendientes_count(),
            })

        if Usuario.objects.filter(username=username).exists():
            messages.error(request, f'El usuario "{username}" ya existe.')
            return render(request, 'admin/crear_medico.html', {
                'form_data': request.POST,
                'especialidades': especialidades,
                'citas_pendientes': _admin_citas_pendientes_count(),
            })
        if email and Usuario.objects.filter(email__iexact=email).exists():
            messages.error(request, f'El correo "{email}" ya está registrado.')
            return render(request, 'admin/crear_medico.html', {
                'form_data': request.POST,
                'especialidades': especialidades,
                'citas_pendientes': _admin_citas_pendientes_count(),
            })

        try:
            with transaction.atomic():
                user = Usuario.objects.create_user(
                    username=username, password=password,
                    first_name=first_name, last_name=last_name,
                    email=email, rol='medico'
                )

                _crear_perfil_medico_compatible(user, especialidad_obj, telefono)
        except Exception as exc:
            logger.exception("Error al crear médico desde el panel admin")
            messages.error(request, 'No se pudo crear el médico. Revisa los datos e intenta nuevamente.')
            return render(request, 'admin/crear_medico.html', {
                'form_data': request.POST,
                'especialidades': especialidades,
                'citas_pendientes': _admin_citas_pendientes_count(),
            })
        messages.success(request, f'Médico "{first_name} {last_name}" creado correctamente.')
        registrar_log(request, 'CREATE', 'Médicos',
            f'Médico "{first_name} {last_name}" (usuario: {username}) registrado', 'INFO')
        return redirect('lista_medicos')

    return render(request, 'admin/crear_medico.html', {
        'especialidades': especialidades,
        'citas_pendientes': _admin_citas_pendientes_count(),
    })


@login_required
def admin_editar_medico(request, medico_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    from medicos.models import Medico
    from landing.models import Especialidad

    medico = get_object_or_404(Medico, id=medico_id)
    especialidades = Especialidad.objects.filter(activo=True)

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        password = request.POST.get('password', '').strip()
        especialidad_id = request.POST.get('especialidad_id', '').strip()
        errores = _validar_usuario_admin(
            first_name,
            last_name,
            medico.usuario.username,
            email,
            password or 'Aa1!aaaa'
        )
        especialidad_obj = None

        if telefono and (not telefono.isdigit() or len(telefono) != 10):
            errores.append("El teléfono debe tener 10 dígitos.")
        if not especialidad_id:
            errores.append("Selecciona una especialidad médica.")
        else:
            try:
                especialidad_obj = especialidades.get(id=especialidad_id)
            except Especialidad.DoesNotExist:
                errores.append("Selecciona una especialidad médica válida.")

        if errores:
            for error in errores:
                messages.error(request, error)
            return render(request, 'admin/editar_medico.html', {
                'medico': medico,
                'especialidades': especialidades,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })

        if email and Usuario.objects.filter(email__iexact=email).exclude(id=medico.usuario.id).exists():
            messages.error(request, f'El correo "{email}" ya está registrado.')
            return render(request, 'admin/editar_medico.html', {
                'medico': medico,
                'especialidades': especialidades,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })

        with transaction.atomic():
            medico.usuario.first_name = first_name
            medico.usuario.last_name  = last_name
            medico.usuario.email      = email
            if password:
                medico.usuario.set_password(password)
            medico.usuario.save()

            medico.especialidad = especialidad_obj
            medico.telefono = telefono
            medico.save()

        messages.success(request, 'Médico actualizado correctamente.')
        registrar_log(request, 'UPDATE', 'Médicos',
            f'Datos del médico "{medico.usuario.get_full_name()}" (ID {medico.id}) actualizados', 'INFO')
        return redirect('lista_medicos')

    return render(request, 'admin/editar_medico.html', {
        'medico': medico,
        'especialidades': especialidades,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


@login_required
def admin_eliminar_medico(request, medico_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    from medicos.models import Medico
    medico = get_object_or_404(Medico, id=medico_id)

    if request.method == 'POST':
        nombre = str(medico)
        medico.usuario.delete()
        messages.success(request, f'Médico "{nombre}" eliminado correctamente.')
        registrar_log(request, 'DELETE', 'Médicos',
            f'Médico "{nombre}" (ID {medico_id}) eliminado del sistema', 'WARNING')
        return redirect('lista_medicos')

    return render(request, 'admin/confirmar_eliminar.html', {
        'objeto': f'Dr. {medico.usuario.first_name} {medico.usuario.last_name}',
        'tipo': 'médico',
        'volver': 'lista_medicos',
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


# ── CITAS ─────────────────────────────────────

@login_required
def admin_eliminar_cita(request, cita_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    cita = get_object_or_404(Cita, id=cita_id)

    if request.method == 'POST':
        cita.delete()
        messages.success(request, 'Cita eliminada correctamente.')
        registrar_log(request, 'DELETE', 'Citas',
            f'Cita ID {cita_id} eliminada', 'WARNING')
        return redirect('todas_citas')

    return render(request, 'admin/confirmar_eliminar.html', {
        'objeto': f'Cita de {cita.paciente.get_full_name()} — {cita.fecha}',
        'tipo': 'cita',
        'volver': 'todas_citas',
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


# ── CONTROLES ─────────────────────────────────

@login_required
def admin_eliminar_control(request, control_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    control = get_object_or_404(ControlPrenatal, id=control_id)

    if request.method == 'POST':
        control.delete()
        messages.success(request, 'Control prenatal eliminado correctamente.')
        registrar_log(request, 'DELETE', 'Controles Prenatales',
            f'Control prenatal ID {control_id} eliminado', 'WARNING')
        return redirect('controles_admin')

    return render(request, 'admin/confirmar_eliminar.html', {
        'objeto': f'Control de {control.paciente.get_full_name()} — {control.fecha}',
        'tipo': 'control prenatal',
        'volver': 'controles_admin',
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })

@login_required
@no_cache_view
def admin_crear_paciente_general(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        password   = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        cedula     = request.POST.get('cedula', '').strip()
        telefono   = request.POST.get('telefono', '').strip()
        edad       = request.POST.get('edad') or None
        direccion  = request.POST.get('direccion', '').strip()
        genero     = request.POST.get('genero', '').strip()
        edad, edad_error = _validar_entero_opcional(edad, 'La edad', 1, 120)

        errores = _validar_usuario_admin(first_name, last_name, username, email, password, password_confirm=password_confirm)
        if genero not in ('femenino', 'masculino', 'otro'):
            errores.append('Selecciona un género válido.')
        if cedula and (not cedula.isdigit() or len(cedula) != 10):
            errores.append('La cédula debe tener exactamente 10 números.')
        if telefono and (not telefono.isdigit() or len(telefono) != 10):
            errores.append('El teléfono debe tener exactamente 10 números.')
        if edad_error:
            errores.append(edad_error)
        if errores:
            for error in errores:
                messages.error(request, error)
            return render(request, 'admin/crear_paciente_general.html', {
                'form_data': request.POST,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })

        if genero not in ('femenino', 'masculino', 'otro'):
            messages.error(request, 'Selecciona un género válido.')
            return render(request, 'admin/crear_paciente_general.html', {
                'form_data': request.POST,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })

        from pacientes.models import Paciente

        if Usuario.objects.filter(username=username).exists():
            messages.error(request, f'El usuario "{username}" ya existe.')
            return render(request, 'admin/crear_paciente_general.html', {
                'form_data': request.POST,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })
        if email and Usuario.objects.filter(email__iexact=email).exists():
            messages.error(request, f'El correo "{email}" ya está registrado.')
            return render(request, 'admin/crear_paciente_general.html', {
                'form_data': request.POST,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })
        if cedula and Paciente.objects.filter(cedula=cedula).exists():
            messages.error(request, f'La cédula "{cedula}" ya está registrada.')
            return render(request, 'admin/crear_paciente_general.html', {
                'form_data': request.POST,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })

        try:
            with transaction.atomic():
                user = Usuario.objects.create_user(
                    username=username, password=password,
                    first_name=first_name, last_name=last_name,
                    email=email, rol='paciente', genero=genero
                )
                paciente, _ = Paciente.objects.get_or_create(usuario=user)
                paciente.cedula = cedula
                paciente.telefono = telefono
                paciente.edad = edad
                paciente.direccion = direccion
                _aplicar_tipo_paciente(paciente, 'general', user)
        except Exception:
            logger.exception("Error al crear paciente general desde el panel admin")
            messages.error(request, 'No se pudo crear el paciente general. Revisa los datos e intenta nuevamente.')
            return render(request, 'admin/crear_paciente_general.html', {
                'form_data': request.POST,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })
        registrar_log(request, 'CREATE', 'Pacientes',
            f'Paciente general "{first_name} {last_name}" registrado', 'INFO')
        messages.success(request, f'Paciente general "{first_name} {last_name}" creado correctamente.')
        return redirect(f"{reverse('lista_pacientes')}?tab=general")

    return render(request, 'admin/crear_paciente_general.html', {
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })

@login_required
@login_required
@no_cache_view
def perfil_enfermera(request):
    rol_lower = request.user.rol.lower() if request.user.rol else ''
    if rol_lower not in ['enfermera', 'secretaria']:
        return redireccionar_por_rol(request.user)
 
    if request.method == 'POST':
        action = request.POST.get('action')
 
        if action == 'perfil':
            request.user.first_name = request.POST.get('first_name', '')
            request.user.last_name  = request.POST.get('last_name', '')
            request.user.email      = request.POST.get('email', '')
            request.user.save()
            messages.success(request, 'Datos actualizados correctamente.')
 
        elif action == 'password':
            from django.contrib.auth import update_session_auth_hash
            password_actual    = request.POST.get('password_actual')
            password_nuevo     = request.POST.get('password_nuevo')
            password_confirmar = request.POST.get('password_confirmar')
 
            if not request.user.check_password(password_actual):
                messages.error(request, 'La contraseña actual es incorrecta.')
            elif password_nuevo != password_confirmar:
                messages.error(request, 'Las contraseñas nuevas no coinciden.')
            elif len(password_nuevo) < 8:
                messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
            else:
                request.user.set_password(password_nuevo)
                request.user.save()
                update_session_auth_hash(request, request.user)
                messages.success(request, 'Contraseña actualizada correctamente.')
 
        return redirect('perfil_enfermera')
 
    return render(request, 'enfermera/perfil_enfermera.html')


# ═══════════════════════════════════════════════════════════════════════════
# CONSULTA GENERAL — Médico general registra consultas a pacientes generales
# ═══════════════════════════════════════════════════════════════════════════

@login_required
@no_cache_view
def registrar_consulta_general(request):
    """Registra una nueva consulta general — todos los médicos pueden usarla."""
    from paciente_general.models import ConsultaGeneral
    from citas.models import Cita
    from pacientes.models import Paciente

    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    # Médico prenatal ve sus pacientes prenatales + pacientes con citas
    # Médico general ve solo sus pacientes con citas
    ids_citas = Cita.objects.filter(
        medico=request.user
    ).values_list('paciente_id', flat=True)

    try:
        es_prenatal = request.user.medico.especialidad and \
                      request.user.medico.especialidad.tipo == 'prenatal'
    except Exception:
        es_prenatal = False

    if es_prenatal:
        # Incluye pacientes con citas + pacientes que este médico tiene asignados como prenatal
        from django.db.models import Q
        pacientes_generales = Paciente.objects.filter(
            Q(usuario_id__in=ids_citas) | Q(medico_prenatal=request.user)
        ).select_related('usuario').distinct().order_by('usuario__first_name', 'usuario__last_name')
    else:
        pacientes_generales = Paciente.objects.filter(
            usuario_id__in=ids_citas
        ).select_related('usuario').order_by('usuario__first_name', 'usuario__last_name')

    # Pre-selección de paciente via ?pac=ID
    pac_preseleccionado = request.GET.get('pac') or request.POST.get('paciente')

    if request.method == 'POST':
        paciente_id = request.POST.get('paciente')
        if not paciente_id:
            messages.error(request, 'Debes seleccionar una paciente.')
            return render(request, 'medico/registrar_consulta_general.html',
                          {'pacientes': pacientes_generales, 'pac_preseleccionado': pac_preseleccionado})

        try:
            paciente_obj = Usuario.objects.get(id=paciente_id, rol='paciente')
        except Usuario.DoesNotExist:
            messages.error(request, 'Paciente no encontrada.')
            return render(request, 'medico/registrar_consulta_general.html',
                          {'pacientes': pacientes_generales, 'pac_preseleccionado': pac_preseleccionado})

        def _val(campo, default=''):
            v = request.POST.get(campo, default).strip()
            return v if v else default

        def _int(campo):
            try: return int(request.POST.get(campo, '')) or None
            except: return None

        def _float(campo):
            try: return float(request.POST.get(campo, '').replace(',', '.')) or None
            except: return None

        def _date(campo):
            from django.utils.dateparse import parse_date
            return parse_date(request.POST.get(campo, '').strip() or '') or None

        def _time(campo):
            from django.utils.dateparse import parse_time
            return parse_time(request.POST.get(campo, '').strip() or '') or None

        try:
            consulta = ConsultaGeneral.objects.create(
            paciente                 = paciente_obj,
            medico                   = request.user,
            especialidad             = _val('especialidad', 'medicina_general'),
            motivo_consulta          = _val('motivo_consulta', '(sin motivo)'),
            antecedentes_personales  = _val('antecedentes_personales'),
            antecedentes_familiares  = _val('antecedentes_familiares'),
            antecedentes_alergicos   = _val('antecedentes_alergicos'),
            antecedentes_quirurgicos = _val('antecedentes_quirurgicos'),
            antecedentes_obstetricos = _val('antecedentes_obstetricos'),
            examen_fisico            = _val('examen_fisico'),
            presion_arterial         = _val('presion_arterial'),
            saturacion_oxigeno       = _int('saturacion_oxigeno'),
            frecuencia_cardiaca      = _int('frecuencia_cardiaca'),
            frecuencia_respiratoria  = _int('frecuencia_respiratoria'),
            temperatura              = _float('temperatura'),
            talla                    = _float('talla'),
            peso                     = _float('peso'),
            diagnostico_clinico      = _val('diagnostico_clinico'),
            procedimiento_realizado  = _val('procedimiento_realizado'),
            hallazgos                = _val('hallazgos'),
            medicamentos_recetados   = _val('medicamentos_recetados'),
            recomendaciones          = _val('recomendaciones'),
            examenes_enviados        = _val('examenes_enviados'),
            evolucion_enfermedad     = _val('evolucion_enfermedad'),
            plan                     = _val('plan'),
            tratamiento              = _val('tratamiento'),
            piezas_dentales_tratadas = _val('piezas_dentales_tratadas'),
            odontograma              = _val('odontograma'),
            tipo_ecografia           = _val('tipo_ecografia'),
            region_examinada         = _val('region_examinada'),
            conclusion_diagnostica   = _val('conclusion_diagnostica'),
            diagnostico_1_patologia  = _val('diag1_patologia'),
            diagnostico_1_cie10      = _val('diag1_cie10'),
            diagnostico_1_presuntivo = bool(request.POST.get('diag1_presuntivo')),
            diagnostico_1_definitivo = bool(request.POST.get('diag1_definitivo')),
            diagnostico_2_patologia  = _val('diag2_patologia'),
            diagnostico_2_cie10      = _val('diag2_cie10'),
            diagnostico_2_presuntivo = bool(request.POST.get('diag2_presuntivo')),
            diagnostico_2_definitivo = bool(request.POST.get('diag2_definitivo')),
            diagnostico_3_patologia  = _val('diag3_patologia'),
            diagnostico_3_cie10      = _val('diag3_cie10'),
            diagnostico_3_presuntivo = bool(request.POST.get('diag3_presuntivo')),
            diagnostico_3_definitivo = bool(request.POST.get('diag3_definitivo')),
            proxima_cita             = _date('proxima_cita'),
            proxima_cita_hora        = _time('proxima_cita_hora'),
            )
        except Exception as e:
            logger.error(f"Error en registrar_consulta_general: {e}", exc_info=True)
            messages.error(request, 'No se pudo registrar la consulta. Revisa los campos e intenta nuevamente.')
            return render(request, 'medico/registrar_consulta_unificada.html',
                          {'pacientes': pacientes_generales, 'pac_preseleccionado': pac_preseleccionado})
        registrar_log(request, 'CREATE', 'Consultas Generales',
            f'Consulta {consulta.get_especialidad_display()} registrada para paciente {paciente_obj.get_full_name()} por {request.user.rol}', 'INFO')
        messages.success(request, f'Consulta de {consulta.get_especialidad_display()} para {paciente_obj.get_full_name()} registrada correctamente.')
        return redirect('ver_consulta_general', consulta_id=consulta.id)

    return render(request, 'medico/registrar_consulta_unificada.html',
                  {'pacientes': pacientes_generales, 'pac_preseleccionado': pac_preseleccionado})


@login_required
@no_cache_view
def historial_consultas_generales(request):
    """Lista las consultas generales. El médico prenatal puede ver las de sus pacientes."""
    from paciente_general.models import ConsultaGeneral
    import json

    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    try:
        es_prenatal = request.user.medico.especialidad and \
                      request.user.medico.especialidad.tipo == 'prenatal'
    except Exception:
        es_prenatal = False

    paciente_id = request.GET.get('paciente')

    if es_prenatal:
        # Médico prenatal ve consultas de todos los médicos para sus pacientes
        from pacientes.models import Paciente
        from django.db.models import Q
        from citas.models import Cita
        ids_citas = Cita.objects.filter(medico=request.user).values_list('paciente_id', flat=True)
        ids_prenatales = Paciente.objects.filter(medico_prenatal=request.user).values_list('usuario_id', flat=True)
        consultas = ConsultaGeneral.objects.filter(
            Q(paciente_id__in=ids_citas) | Q(paciente_id__in=ids_prenatales)
        ).select_related('paciente', 'medico').order_by('-fecha')
    else:
        # Médico general solo ve sus propias consultas
        consultas = ConsultaGeneral.objects.filter(
            medico=request.user
        ).select_related('paciente', 'medico').order_by('-fecha')

    if paciente_id:
        consultas = consultas.filter(paciente_id=paciente_id)

    # Selector de pacientes
    from citas.models import Cita
    ids_mis_pacientes = Cita.objects.filter(
        medico=request.user
    ).values_list('paciente_id', flat=True)

    from pacientes.models import Paciente
    if es_prenatal:
        from django.db.models import Q
        ids_prenatales = Paciente.objects.filter(medico_prenatal=request.user).values_list('usuario_id', flat=True)
        pacientes_generales = Paciente.objects.filter(
            Q(usuario_id__in=ids_mis_pacientes) | Q(usuario_id__in=ids_prenatales)
        ).select_related('usuario').distinct().order_by('usuario__first_name')
    else:
        pacientes_generales = Paciente.objects.filter(
            usuario_id__in=ids_mis_pacientes
        ).select_related('usuario').order_by('usuario__first_name')

    # JSON para búsqueda en tiempo real
    pacientes_json = json.dumps([{
        'id': p.usuario.id,
        'nombre': p.usuario.get_full_name() or p.usuario.username,
        'cedula': p.usuario.username,
    } for p in pacientes_generales])

    return render(request, 'medico/historial_consultas_generales.html', {
        'consultas': consultas,
        'pacientes': pacientes_generales,
        'pacientes_json': pacientes_json,
        'paciente_filtrado_id': int(paciente_id) if paciente_id else None,
    })


@login_required
@no_cache_view
def ver_consulta_general(request, consulta_id):
    """Muestra el detalle completo de una consulta general."""
    from paciente_general.models import ConsultaGeneral

    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    consulta = get_object_or_404(ConsultaGeneral, id=consulta_id)
    return render(request, 'medico/ver_consulta_general.html', {'consulta': consulta})


# ═══════════════════════════════════════════════════════════════════════════
# PROGRAMACIÓN DE PARTOS — Médico prenatal programa, paciente prenatal ve
# ═══════════════════════════════════════════════════════════════════════════

@login_required
@no_cache_view
def programar_parto(request):
    """El médico prenatal programa una fecha/hora de parto para una paciente."""
    import json
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        from paciente_general.models import ProgramacionParto
        from pacientes.models import Paciente
        from citas.models import Cita

        if request.user.rol != 'medico':
            return redireccionar_por_rol(request.user)
        
        # Solo pacientes con embarazo activo y vinculadas al médico
        ids_embarazo_activo = Paciente.objects.filter(
            estado_embarazo='ACTIVO',
        ).values_list('usuario_id', flat=True)
        
        ids_asignadas = Paciente.objects.filter(
            medico_prenatal=request.user
        ).values_list('usuario_id', flat=True)
        ids_con_citas = Cita.objects.filter(
            medico=request.user
        ).values_list('paciente_id', flat=True)
        todos_ids = set(ids_asignadas) | set(ids_con_citas)
        
        # Filtrar: solo pacientes del médico Y con embarazo activo
        todos_ids = todos_ids & set(ids_embarazo_activo)

        pacientes_prenatales = Usuario.objects.filter(
            id__in=todos_ids, rol='paciente', is_active=True
        ).distinct().order_by('first_name', 'last_name')

        # Preparar JSON de pacientes para búsqueda en tiempo real
        pacientes_json = json.dumps([{
            'id': p.id,
            'nombre': p.get_full_name() or p.username,
            'cedula': p.username,  # Usar username como identificador
        } for p in pacientes_prenatales])
    except Exception as e:
        logger.error(f"Error en programar_parto (GET): {str(e)}", exc_info=True)
        messages.error(request, f'Error al cargar la página: {str(e)}')
        return redirect('medico_dashboard')

    if request.method == 'POST':
        try:
            paciente_id      = request.POST.get('paciente')
            tipo             = request.POST.get('tipo', 'parto_natural')
            fecha_programada = request.POST.get('fecha_programada')
            hora_programada  = request.POST.get('hora_programada')
            semanas          = request.POST.get('semanas_gestacion') or None
            lugar            = request.POST.get('lugar', 'Zumedical — Centro Médico').strip()
            indicaciones     = request.POST.get('indicaciones', '').strip()

            if not all([paciente_id, fecha_programada, hora_programada]):
                messages.error(request, 'Paciente, fecha y hora son obligatorios.')
                return render(request, 'medico/programar_parto.html',
                              {'pacientes': pacientes_prenatales, 'pacientes_json': pacientes_json})

            try:
                paciente_obj = Usuario.objects.filter(
                    id=paciente_id, rol='paciente', id__in=todos_ids
                ).distinct().get()
                
                # Verificar que la paciente tenga embarazo activo
                paciente_perfil = Paciente.objects.get(usuario=paciente_obj)
                if paciente_perfil.estado_embarazo != 'ACTIVO':
                    messages.error(request, 'La paciente debe tener un embarazo activo para programar el parto.')
                    return render(request, 'medico/programar_parto.html',
                                  {'pacientes': pacientes_prenatales, 'pacientes_json': pacientes_json})
                    
            except Usuario.DoesNotExist:
                messages.error(request, 'Paciente prenatal no encontrada.')
                return render(request, 'medico/programar_parto.html',
                              {'pacientes': pacientes_prenatales, 'pacientes_json': pacientes_json})
            except Paciente.DoesNotExist:
                messages.error(request, 'Perfil de paciente no encontrado.')
                return render(request, 'medico/programar_parto.html',
                              {'pacientes': pacientes_prenatales, 'pacientes_json': pacientes_json})

            parto = ProgramacionParto.objects.create(
                paciente          = paciente_obj,
                medico            = request.user,
                tipo              = tipo,
                fecha_programada  = fecha_programada,
                hora_programada   = hora_programada,
                semanas_gestacion = int(semanas) if semanas else None,
                lugar             = lugar or 'Zumedical — Centro Médico',
                indicaciones      = indicaciones,
                estado            = 'programado',
            )
            registrar_log(request, 'CREATE', 'Programación de Partos',
                f'{parto.get_tipo_display()} programado para {paciente_obj.get_full_name()} - Fecha: {fecha_programada} {hora_programada}', 'INFO')
            # Parsear las fechas que vienen como strings desde el POST
            from datetime import datetime
            try:
                f_str = datetime.strptime(str(fecha_programada), "%Y-%m-%d").strftime("%d/%m/%Y")
            except:
                f_str = str(fecha_programada)
            
            try:
                h_str = datetime.strptime(str(hora_programada), "%H:%M").strftime("%H:%M")
            except:
                h_str = str(hora_programada)[:5]

            messages.success(
                request,
                f'{parto.get_tipo_display()} programado para {paciente_obj.get_full_name()} '
                f'el {f_str} a las {h_str}.'
            )
            return redirect('lista_programaciones_parto')
        except Exception as e:
            logger.error(f"Error en programar_parto (POST): {str(e)}", exc_info=True)
            messages.error(request, f'Error al guardar la programación: {str(e)}')
            return render(request, 'medico/programar_parto.html',
                          {'pacientes': pacientes_prenatales, 'pacientes_json': pacientes_json})

    return render(request, 'medico/programar_parto.html',
                  {'pacientes': pacientes_prenatales, 'pacientes_json': pacientes_json})


@login_required
@no_cache_view
def editar_parto(request, parto_id):
    """Edita o cambia el estado de una programación de parto."""
    import json
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        from paciente_general.models import ProgramacionParto

        if request.user.rol != 'medico':
            return redireccionar_por_rol(request.user)

        parto = get_object_or_404(ProgramacionParto, id=parto_id)

        if request.method == 'POST':
            try:
                parto.tipo             = request.POST.get('tipo', parto.tipo)
                parto.fecha_programada = request.POST.get('fecha_programada', str(parto.fecha_programada))
                parto.hora_programada  = request.POST.get('hora_programada', str(parto.hora_programada))
                parto.lugar            = request.POST.get('lugar', parto.lugar).strip() or parto.lugar
                parto.indicaciones     = request.POST.get('indicaciones', parto.indicaciones).strip()
                nuevo_estado           = request.POST.get('estado', parto.estado)
                parto.estado           = nuevo_estado
                sem = request.POST.get('semanas_gestacion')
                if sem:
                    try: parto.semanas_gestacion = int(sem)
                    except: pass
                parto.save()
                registrar_log(request, 'UPDATE', 'Programación de Partos',
                    f'Programación de parto #{parto.id} actualizada - Nuevo estado: {nuevo_estado}', 'INFO')
                messages.success(request, 'Programación de parto actualizada.')
                return redirect('lista_programaciones_parto')
            except Exception as e:
                logger.error(f"Error en editar_parto (POST): {str(e)}", exc_info=True)
                messages.error(request, f'Error al actualizar la programación: {str(e)}')

        pacientes_prenatales = Usuario.objects.filter(
            rol='paciente', is_active=True, paciente__estado_embarazo='ACTIVO'
        ).distinct().order_by('first_name')
        
        # Preparar JSON de pacientes para búsqueda en tiempo real
        pacientes_json = json.dumps([{
            'id': p.id,
            'nombre': p.get_full_name() or p.username,
            'cedula': p.username,
        } for p in pacientes_prenatales])
        
        return render(request, 'medico/programar_parto.html', {
            'pacientes': pacientes_prenatales,
            'pacientes_json': pacientes_json,
            'parto':     parto,
            'modo':      'editar',
        })
    except Exception as e:
        logger.error(f"Error en editar_parto (GET): {str(e)}", exc_info=True)
        messages.error(request, f'Error al cargar la programación: {str(e)}')
        return redirect('lista_programaciones_parto')


@login_required
@no_cache_view
def eliminar_parto(request, parto_id):
    """Elimina una programación de parto desde la lista del médico."""
    import logging
    logger = logging.getLogger(__name__)
    
    logger.info(f"Intentando eliminar parto con ID: {parto_id}")
    logger.info(f"Usuario: {request.user.username}, Método: {request.method}")
    
    try:
        from paciente_general.models import ProgramacionParto

        if request.user.rol != 'medico':
            messages.error(request, 'Solo los médicos pueden eliminar programaciones de parto.')
            return redireccionar_por_rol(request.user)

        # Buscar la programación
        try:
            parto = ProgramacionParto.objects.get(id=parto_id)
            logger.info(f"Parto encontrado: {parto}")
        except ProgramacionParto.DoesNotExist:
            logger.error(f"No se encontró ProgramacionParto con ID {parto_id}")
            # Listar todas las programaciones para debug
            todas = ProgramacionParto.objects.all().values_list('id', flat=True)
            logger.error(f"IDs existentes: {list(todas)}")
            messages.error(request, f'No se encontró la programación de parto con ID {parto_id}.')
            return redirect('lista_programaciones_parto')
        
        paciente_nombre = parto.paciente.get_full_name() or parto.paciente.username

        if request.method == 'POST':
            parto.delete()
            registrar_log(request, 'DELETE', 'Programación de Partos',
                f'Programación de parto #{parto_id} eliminada para {paciente_nombre}', 'WARNING')
            messages.success(request, f'Programación de parto de {paciente_nombre} eliminada correctamente.')
            return redirect('lista_programaciones_parto')

        messages.error(request, 'Debe usar el método POST para eliminar la programación.')
        return redirect('lista_programaciones_parto')
        
    except Exception as e:
        logger.error(f"Error en eliminar_parto: {str(e)}", exc_info=True)
        messages.error(request, f'Error al eliminar la programación: {str(e)}')
        return redirect('lista_programaciones_parto')


@login_required
@no_cache_view
def lista_programaciones_parto(request):
    """Lista todas las programaciones de parto del médico."""
    import logging
    from collections import OrderedDict
    from django.utils import timezone
    from datetime import datetime
    
    logger = logging.getLogger(__name__)
    
    try:
        from paciente_general.models import ProgramacionParto

        if request.user.rol != 'medico':
            return redireccionar_por_rol(request.user)

        ahora = timezone.localtime()
        hoy = timezone.localdate()
        ProgramacionParto.objects.filter(
            medico=request.user,
            estado__in=['programado', 'confirmado', 'reprogramado'],
        ).filter(
            Q(fecha_programada__lt=hoy) |
            Q(fecha_programada=hoy, hora_programada__lte=ahora.time())
        ).update(estado='realizado')

        busqueda = request.GET.get('q', '').strip()
        fecha_desde = request.GET.get('fecha_desde', '').strip()
        fecha_hasta = request.GET.get('fecha_hasta', '').strip()

        programaciones_qs = ProgramacionParto.objects.select_related(
            'paciente', 'medico'
        ).filter(medico=request.user)

        if busqueda:
            programaciones_qs = programaciones_qs.filter(
                Q(paciente__first_name__icontains=busqueda) |
                Q(paciente__last_name__icontains=busqueda) |
                Q(paciente__username__icontains=busqueda) |
                Q(paciente__paciente__cedula__icontains=busqueda)
            )
        if fecha_desde:
            programaciones_qs = programaciones_qs.filter(fecha_programada__gte=fecha_desde)
        if fecha_hasta:
            programaciones_qs = programaciones_qs.filter(fecha_programada__lte=fecha_hasta)

        programaciones = list(programaciones_qs.order_by('-fecha_programada', 'hora_programada'))

        def agrupar_por_mes(items):
            grupos = OrderedDict()
            for item in items:
                if item.fecha_programada:
                    clave = item.fecha_programada.strftime('%Y-%m')
                    mes = item.fecha_programada.strftime('%B %Y').capitalize()
                else:
                    clave = 'sin-fecha'
                    mes = 'Sin fecha'
                if clave not in grupos:
                    grupos[clave] = {'mes': mes, 'items': []}
                grupos[clave]['items'].append(item)
            return list(grupos.values())

        activas, historial = [], []
        for p in programaciones:
            fecha_hora = None
            if p.fecha_programada and p.hora_programada:
                fecha_hora = timezone.make_aware(
                    datetime.combine(p.fecha_programada, p.hora_programada),
                    timezone.get_current_timezone()
                )
            p.ya_paso = bool(fecha_hora and fecha_hora < ahora)
            p.pendiente_cierre = p.ya_paso and p.estado in ('programado', 'confirmado', 'reprogramado')
            perfil = getattr(p.paciente, 'paciente', None)
            p.prenatal_activo = bool(
                perfil
                and getattr(perfil, 'estado_embarazo', '') == 'ACTIVO'
            )
            if p.estado in ('realizado', 'cancelado') or p.ya_paso or not p.prenatal_activo:
                historial.append(p)
            else:
                activas.append(p)

        return render(request, 'medico/lista_programaciones_parto.html',
                      {
                          'programaciones': programaciones,
                          'activas': activas,
                          'historial': historial,
                          'grupos_activas': agrupar_por_mes(activas),
                          'grupos_historial': agrupar_por_mes(historial),
                          'filtros': {
                              'q': busqueda,
                              'fecha_desde': fecha_desde,
                              'fecha_hasta': fecha_hasta,
                          },
                      })
    except Exception as e:
        logger.error(f"Error en lista_programaciones_parto: {str(e)}", exc_info=True)
        messages.error(request, f'Error al cargar las programaciones: {str(e)}')
        return redirect('medico_dashboard')


@login_required
@no_cache_view
def perfil_medico(request):
    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'perfil':
            request.user.first_name = request.POST.get('first_name', '')
            request.user.last_name  = request.POST.get('last_name', '')
            request.user.email      = request.POST.get('email', '')
            request.user.save()
            messages.success(request, 'Datos actualizados correctamente.')

        elif action == 'password':
            actual    = request.POST.get('password_actual')
            nueva     = request.POST.get('password_nuevo')
            confirmar = request.POST.get('password_confirmar')

            if not request.user.check_password(actual):
                messages.error(request, 'La contraseña actual es incorrecta.')
            elif nueva != confirmar:
                messages.error(request, 'Las contraseñas nuevas no coinciden.')
            elif len(nueva) < 8:
                messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
            else:
                request.user.set_password(nueva)
                request.user.save()
                update_session_auth_hash(request, request.user)
                messages.success(request, 'Contraseña actualizada correctamente.')

        return redirect('perfil_medico')

    try:
        es_prenatal = request.user.medico.especialidad and \
                      request.user.medico.especialidad.tipo == 'prenatal'
    except:
        es_prenatal = True

    if es_prenatal:
        return render(request, 'medico/perfil_medico.html', {'es_prenatal': es_prenatal})
    else:
        return render(request, 'medico/perfil_medico_general.html', {'es_prenatal': es_prenatal})


# ── AUDITORÍA ──────────────────────────────────────────────────

@login_required
@no_cache_view
def auditoria_admin(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    hoy = timezone.now().date()

    # Queryset base
    logs_qs = LogAuditoria.objects.select_related('usuario').all()

    # Filtros GET
    accion    = request.GET.get('accion', '')
    severidad = request.GET.get('severidad', '')
    usuario   = request.GET.get('usuario', '')
    desde     = request.GET.get('desde', '')
    hasta     = request.GET.get('hasta', '')

    if accion:
        logs_qs = logs_qs.filter(accion=accion)
    if severidad:
        logs_qs = logs_qs.filter(severidad=severidad)
    if usuario:
        logs_qs = logs_qs.filter(usuario__id=usuario)
    if desde:
        logs_qs = logs_qs.filter(fecha__date__gte=desde)
    if hasta:
        logs_qs = logs_qs.filter(fecha__date__lte=hasta)

    # Exportar Excel
    if request.GET.get('export') == 'csv':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse(
                'El módulo openpyxl no está instalado. Ejecuta: pip install openpyxl',
                status=500, content_type='text/plain'
            )
        from django.utils import timezone as tz
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Auditoría Zumedical"

        # Título principal
        ws.merge_cells('A1:H1')
        titulo = ws['A1']
        titulo.value = f'Reporte de Auditoría — Zumedical — {hoy.strftime("%d/%m/%Y")}'
        titulo.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
        titulo.fill = PatternFill('solid', fgColor='8A2563')
        titulo.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Encabezados
        headers = ['N°', 'Fecha', 'Hora', 'Usuario', 'Rol', 'Acción', 'Módulo', 'IP', 'Severidad', 'Descripción']
        ws.append([])  # fila vacía
        ws.append(headers)
        header_row = ws.max_row
        header_fill = PatternFill('solid', fgColor='CC4D99')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        thin = Side(style='thin', color='E8AAD4')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[header_row].height = 22

        # Datos
        fill_par = PatternFill('solid', fgColor='FBF0F7')
        fill_imp = PatternFill('solid', fgColor='FFFFFF')
        font_data = Font(name='Calibri', size=10)

        for i, log in enumerate(logs_qs.order_by('-fecha'), 1):
            fecha_local = tz.localtime(log.fecha)
            nombre = log.usuario.get_full_name() if log.usuario else '—'
            rol = log.usuario.rol.capitalize() if log.usuario else '—'
            row = [
                i,
                fecha_local.strftime('%d/%m/%Y'),
                fecha_local.strftime('%H:%M:%S'),
                nombre,
                rol,
                log.get_accion_display(),
                log.modulo or '—',
                log.ip_address or '—',
                log.get_severidad_display(),
                log.descripcion or '—',
            ]
            ws.append(row)
            current_row = ws.max_row
            fill = fill_par if i % 2 == 0 else fill_imp
            for col_num in range(1, len(row) + 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.font = font_data
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=col_num == len(row))

        # Anchos de columna
        anchos = [6, 12, 10, 22, 12, 16, 16, 14, 12, 50]
        for col_num, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = ancho

        # Guardar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="auditoria_zumedical_{hoy}.xlsx"'
        return response

    # KPIs — calcular rango UTC del día local en Ecuador
    from django.utils import timezone as tz
    from zoneinfo import ZoneInfo
    from datetime import datetime, timedelta

    ecuador = ZoneInfo('America/Guayaquil')
    ahora_ecuador = tz.now().astimezone(ecuador)
    hoy_local = ahora_ecuador.date()

    # Convertir inicio y fin del día Ecuador a UTC para filtrar correctamente
    inicio_dia_ec = datetime(hoy_local.year, hoy_local.month, hoy_local.day, 0, 0, 0, tzinfo=ecuador)
    fin_dia_ec    = datetime(hoy_local.year, hoy_local.month, hoy_local.day, 23, 59, 59, tzinfo=ecuador)
    inicio_utc = inicio_dia_ec.astimezone(ZoneInfo('UTC'))
    fin_utc    = fin_dia_ec.astimezone(ZoneInfo('UTC'))

    total_logs           = LogAuditoria.objects.count()
    logs_criticos_hoy    = LogAuditoria.objects.filter(
        fecha__gte=inicio_utc, fecha__lte=fin_utc, severidad='CRITICAL').count()
    logins_hoy           = LogAuditoria.objects.filter(
        fecha__gte=inicio_utc, fecha__lte=fin_utc, accion='LOGIN').count()
    usuarios_activos_hoy = LogAuditoria.objects.filter(
        fecha__gte=inicio_utc, fecha__lte=fin_utc).values('usuario').distinct().count()
    
    # Paginación: 25 registros por página
    paginator = Paginator(logs_qs, 25)
    page_num  = request.GET.get('page', 1)
    logs_page = paginator.get_page(page_num)

    return render(request, 'admin/auditoria_admin.html', {
        'logs':                  logs_page,
        'total_logs':            total_logs,
        'logs_criticos_hoy':     logs_criticos_hoy,
        'logins_hoy':            logins_hoy,
        'usuarios_activos_hoy':  usuarios_activos_hoy,
        'usuarios_lista':        User.objects.filter(is_active=True).order_by('first_name'),
        'citas_pendientes':      Cita.objects.filter(estado='pendiente').count(),
        'config_retencion':      request.session.get('auditoria_retencion', '90'),
    })


@login_required
@no_cache_view
def auditoria_admin_data(request):
    """
    Endpoint JSON para el panel de Auditoria.
    Devuelve KPIs, datos para graficos (actividad 7 dias + dona por accion)
    y filas de tabla paginadas, respetando los mismos filtros GET que la
    vista principal. Usado para busqueda en tiempo real, filtros sin
    recargar, auto-refresh y los graficos.
    """
    if request.user.rol != 'admin':
        return JsonResponse({'error': 'No autorizado'}, status=403)

    hoy = timezone.now().date()

    logs_qs = LogAuditoria.objects.select_related('usuario').all()

    accion    = request.GET.get('accion', '')
    severidad = request.GET.get('severidad', '')
    usuario   = request.GET.get('usuario', '')
    desde     = request.GET.get('desde', '')
    hasta     = request.GET.get('hasta', '')
    q         = request.GET.get('q', '').strip()

    if accion:
        logs_qs = logs_qs.filter(accion=accion)
    if severidad:
        logs_qs = logs_qs.filter(severidad=severidad)
    if usuario:
        logs_qs = logs_qs.filter(usuario__id=usuario)
    if desde:
        logs_qs = logs_qs.filter(fecha__date__gte=desde)
    if hasta:
        logs_qs = logs_qs.filter(fecha__date__lte=hasta)
    if q:
        logs_qs = logs_qs.filter(
            Q(usuario__first_name__icontains=q) |
            Q(usuario__last_name__icontains=q) |
            Q(usuario__username__icontains=q) |
            Q(modulo__icontains=q) |
            Q(descripcion__icontains=q) |
            Q(ip_address__icontains=q)
        )

    from zoneinfo import ZoneInfo
    from datetime import datetime
    ecuador = ZoneInfo('America/Guayaquil')
    ahora_ec = timezone.now().astimezone(ecuador)
    hoy_ec = ahora_ec.date()

    # Rango UTC del día local en Ecuador
    inicio_dia_ec = datetime(hoy_ec.year, hoy_ec.month, hoy_ec.day, 0, 0, 0, tzinfo=ecuador)
    fin_dia_ec    = datetime(hoy_ec.year, hoy_ec.month, hoy_ec.day, 23, 59, 59, tzinfo=ecuador)
    inicio_utc = inicio_dia_ec.astimezone(ZoneInfo('UTC'))
    fin_utc    = fin_dia_ec.astimezone(ZoneInfo('UTC'))

    total_logs           = LogAuditoria.objects.count()
    logs_criticos_hoy    = LogAuditoria.objects.filter(fecha__gte=inicio_utc, fecha__lte=fin_utc, severidad='CRITICAL').count()
    logins_hoy           = LogAuditoria.objects.filter(fecha__gte=inicio_utc, fecha__lte=fin_utc, accion='LOGIN').count()
    usuarios_activos_hoy = LogAuditoria.objects.filter(fecha__gte=inicio_utc, fecha__lte=fin_utc).values('usuario').distinct().count()

    # Grafico 1: actividad por dia (ultimos 7 dias) — usando fecha Ecuador
    actividad_labels = []
    actividad_data = []
    utc = ZoneInfo('UTC')
    for i in range(6, -1, -1):
        dia = hoy_ec - timedelta(days=i)
        # Convertir el día Ecuador a rango UTC para filtrar correctamente
        inicio = datetime(dia.year, dia.month, dia.day, 0, 0, 0, tzinfo=ecuador).astimezone(utc)
        fin    = datetime(dia.year, dia.month, dia.day, 23, 59, 59, tzinfo=ecuador).astimezone(utc)
        count = LogAuditoria.objects.filter(fecha__range=(inicio, fin)).count()
        actividad_labels.append(dia.strftime('%d/%m'))
        actividad_data.append(count)

    # Grafico 2: dona por tipo de accion (sobre el queryset filtrado)
    acciones_count = (
        logs_qs.values('accion')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    try:
        accion_labels_map = dict(LogAuditoria._meta.get_field('accion').choices)
    except Exception:
        accion_labels_map = {}
    dona_labels = [accion_labels_map.get(a['accion'], a['accion']) for a in acciones_count]
    dona_data   = [a['total'] for a in acciones_count]

    paginator = Paginator(logs_qs, 25)
    page_num  = request.GET.get('page', 1)
    logs_page = paginator.get_page(page_num)

    rows = []
    for log in logs_page:
        rows.append({
            'id':          log.id,
            'fecha':       log.fecha.strftime('%d/%m/%Y'),
            'hora':        log.fecha.strftime('%H:%M:%S'),
            'usuario':     log.usuario.get_full_name() if log.usuario and log.usuario.get_full_name() else (log.usuario.username if log.usuario else '\u2014'),
            'usuario_inicial': (log.usuario.first_name[:1] if log.usuario and log.usuario.first_name else (log.usuario.username[:1] if log.usuario else '?')).upper(),
            'rol':         log.usuario.groups.first().name if log.usuario and log.usuario.groups.first() else 'Usuario',
            'accion':      log.accion,
            'accion_display': log.get_accion_display(),
            'modulo':      log.modulo or '\u2014',
            'ip':          log.ip_address or '\u2014',
            'severidad':   log.severidad,
            'severidad_display': log.get_severidad_display(),
            'descripcion': log.descripcion or 'Sin descripcion',
        })

    return JsonResponse({
        'kpis': {
            'total_logs': total_logs,
            'logs_criticos_hoy': logs_criticos_hoy,
            'logins_hoy': logins_hoy,
            'usuarios_activos_hoy': usuarios_activos_hoy,
        },
        'charts': {
            'actividad_labels': actividad_labels,
            'actividad_data': actividad_data,
            'dona_labels': dona_labels,
            'dona_data': dona_data,
        },
        'rows': rows,
        'pagination': {
            'current_page': logs_page.number,
            'num_pages': paginator.num_pages,
            'count': paginator.count,
            'has_previous': logs_page.has_previous(),
            'has_next': logs_page.has_next(),
        },
    })
@login_required
@no_cache_view
def auditoria_config(request):
    """Guarda la configuración de retención en sesión."""
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)
    if request.method == 'POST':
        dias = request.POST.get('retencion', '90')
        if dias in ['0', '30', '90', '180', '365']:
            request.session['auditoria_retencion'] = dias
            messages.success(request, f'Configuración guardada: {"Nunca eliminar" if dias == "0" else f"{dias} días de retención"}.')
        else:
            messages.error(request, 'Valor de retención inválido.')
    return redirect('auditoria_admin')
 
 
@login_required
@no_cache_view
def auditoria_limpiar(request):
    """Elimina registros de auditoría según la retención configurada."""
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)
    if request.method == 'POST':
        dias = request.session.get('auditoria_retencion', '90')
        if dias == '0':
            messages.warning(request, 'La configuración actual es "Nunca eliminar". No se eliminó ningún registro.')
        else:
            from django.utils import timezone
            from datetime import timedelta
            limite = timezone.now() - timedelta(days=int(dias))
            eliminados = LogAuditoria.objects.filter(fecha__lt=limite).count()
            LogAuditoria.objects.filter(fecha__lt=limite).delete()
            registrar_log(request, 'DELETE', 'Auditoría',
                          f'Eliminados {eliminados} registros con más de {dias} días.', 'WARNING')
            messages.success(request, f'Se eliminaron {eliminados} registros con más de {dias} días.')
    return redirect('auditoria_admin')

# ── ENDPOINTS DE REPORTES ─────────────────────────────────────

@login_required
def reporte_pacientes_data(request):
    if request.user.rol != 'admin':
        return JsonResponse({'error': 'No autorizado'}, status=403)

    from pacientes.models import Paciente
    pacientes = Paciente.objects.select_related('usuario').order_by('usuario__last_name')
    rows = []
    for i, p in enumerate(pacientes, 1):
        rows.append({
            'id':         i,
            'codigo':     f'PAC-{p.id:04d}',
            'cedula':     p.cedula or '—',
            'nombre':     p.usuario.get_full_name() or p.usuario.username,
            'edad':       p.edad if p.edad else '—',
            'telefono':   p.telefono or '—',
            'email':      p.usuario.email or '—',
            'direccion':  p.direccion or '—',
            'registro':   p.usuario.date_joined.strftime('%d/%m/%Y'),
            'estado':     'Activa' if p.usuario.is_active else 'Inactiva',
            'tipo':       'Prenatal' if p.usuario.puede_prenatal else 'General',
        })
    return JsonResponse({'rows': rows})


@login_required
def reporte_medicos_data(request):
    if request.user.rol != 'admin':
        return JsonResponse({'error': 'No autorizado'}, status=403)

    from medicos.models import Medico
    medicos = Medico.objects.select_related('usuario', 'especialidad').order_by('usuario__last_name')
    rows = []
    for i, m in enumerate(medicos, 1):
        rows.append({
            'id':           f'MED-{m.id:04d}',
            'nombre':       m.usuario.get_full_name() or m.usuario.username,
            'especialidad': m.especialidad.nombre if m.especialidad else '—',
            'telefono':     m.telefono or '—',
            'email':        m.usuario.email or '—',
            'estado':       'Activo' if m.usuario.is_active else 'Inactivo',
        })
    return JsonResponse({'rows': rows})


@login_required
def reporte_citas_data(request):
    if request.user.rol != 'admin':
        return JsonResponse({'error': 'No autorizado'}, status=403)

    from citas.models import Cita
    citas = Cita.objects.select_related('paciente', 'medico', 'especialidad').order_by('-fecha', 'hora')
    rows = []
    for i, c in enumerate(citas, 1):
        rows.append({
            'numero':       i,
            'paciente':     c.paciente.get_full_name() or c.paciente.username,
            'medico':       c.medico.get_full_name() or c.medico.username,
            'especialidad': c.especialidad.nombre if c.especialidad else '—',
            'fecha':        c.fecha.strftime('%d/%m/%Y'),
            'hora':         c.hora.strftime('%H:%M'),
            'estado':       c.get_estado_display(),
            'motivo':       c.motivo or '—',
        })
    return JsonResponse({'rows': rows})


@login_required
def reporte_controles_data(request):
    if request.user.rol != 'admin':
        return JsonResponse({'error': 'No autorizado'}, status=403)

    from control_prenatal.models import ControlPrenatal
    controles = ControlPrenatal.objects.select_related('paciente', 'medico').order_by('-fecha')
    rows = []
    for i, c in enumerate(controles, 1):
        rows.append({
            'numero':        i,
            'paciente':      c.paciente.get_full_name() or c.paciente.username,
            'medico':        c.medico.get_full_name() or c.medico.username,
            'fecha':         c.fecha.strftime('%d/%m/%Y'),
            'semanas':       c.semanas_gestacion,
            'peso':          f'{c.peso} kg',
            'presion':       c.presion_arterial or '—',
            'observaciones': c.observaciones or '—',
        })
    return JsonResponse({'rows': rows})


@login_required
def reporte_usuarios_data(request):
    if request.user.rol != 'admin':
        return JsonResponse({'error': 'No autorizado'}, status=403)

    from django.contrib.sessions.models import Session
    from django.utils import timezone as tz

    sesiones_activas = Session.objects.filter(expire_date__gte=tz.now())
    ids_activos = set()
    for s in sesiones_activas:
        uid = s.get_decoded().get('_auth_user_id')
        if uid:
            ids_activos.add(int(uid))

    usuarios = User.objects.exclude(rol='paciente').order_by('rol', 'last_name')
    rows = []
    for i, u in enumerate(usuarios, 1):
        rows.append({
            'id':            f'USR-{u.id:04d}',
            'usuario':       u.username,
            'nombre':        u.get_full_name() or '—',
            'rol':           u.rol.capitalize(),
            'email':         u.email or '—',
            'estado':        'Activo' if u.is_active else 'Inactivo',
            'ultimo_acceso': u.last_login.strftime('%d/%m/%Y %H:%M') if u.last_login else 'Nunca',
            'en_linea':      u.id in ids_activos,
        })
    return JsonResponse({'rows': rows})


@login_required
@no_cache_view
def cambiar_estado_cita(request, cita_id):
    if request.method == 'POST':
        cita = get_object_or_404(Cita, id=cita_id)
        
        # Validar permisos
        rol_lower = request.user.rol.lower() if request.user.rol else ''
        if rol_lower not in ['medico', 'admin', 'administrador', 'enfermera', 'secretaria'] or (rol_lower == 'medico' and cita.medico != request.user):
            messages.error(request, 'No tienes permiso para modificar esta cita.')
            return redirect(request.META.get('HTTP_REFERER', '/'))
            
        nuevo_estado = request.POST.get('estado')
        estados_validos = dict(Cita.ESTADO).keys()
        
        if nuevo_estado in estados_validos:
            cita.estado = nuevo_estado
            
            # Si es cancelación, guardar motivo
            if nuevo_estado in ['cancelado', 'cancelada']:
                motivo_cancelacion = request.POST.get('motivo_cancelacion', '').strip()
                cita.motivo_cancelacion = motivo_cancelacion
                registrar_log(request, 'CANCELACION', 'Citas',
                    f'Cita #{cita.id} cancelada por {request.user.rol}. Motivo: {motivo_cancelacion}', 'WARNING')
            else:
                registrar_log(request, 'UPDATE', 'Citas',
                    f'Cita #{cita.id} estado actualizado a {nuevo_estado} por {request.user.rol}', 'INFO')
            
            cita.save()
            messages.success(request, f'Estado de la cita actualizado a {cita.get_estado_display()}.')
        else:
            messages.error(request, 'Estado seleccionado no es válido.')
            
    return redirect(request.META.get('HTTP_REFERER', 'citas_medico'))




# ═════════════════════════════════════════════════════════════════════════
# GESTIÓN DE PACIENTES - BÚSQUEDA Y GESTIÓN DE EMBARAZOS
# ═════════════════════════════════════════════════════════════════════════

@login_required
def buscar_paciente(request):
    """Vista para buscar paciente por cédula (Admin/Secretaria)"""
    if request.user.rol not in ['admin', 'secretaria', 'enfermera']:
        return redireccionar_por_rol(request.user)

    paciente = None
    cédula = request.GET.get('cedula', '').strip()
    error_msg = None

    if cédula:
        from pacientes.models import Paciente
        try:
            paciente = Paciente.objects.get(cedula=cédula)
        except Paciente.DoesNotExist:
            error_msg = f"Paciente con cédula {cédula} no encontrado"

    context = {
        'paciente': paciente,
        'cedula_buscada': cédula,
        'error_msg': error_msg,
    }
    return render(request, 'admin/buscar_paciente.html', context)


@login_required
def ficha_paciente(request, paciente_id):
    """Vista para ver ficha completa del paciente (Admin/Secretaria)"""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        if request.user.rol not in ['admin', 'secretaria', 'enfermera', 'medico']:
            return redireccionar_por_rol(request.user)

        from pacientes.models import Paciente, Embarazo
        from paciente_general.models import ConsultaGeneral

        paciente = get_object_or_404(Paciente, id=paciente_id)
        
        # Datos del paciente
        try:
            embarazos = paciente.embarazos.all().order_by('-fecha_inicio')
        except Exception as e:
            logger.error(f"Error al obtener embarazos: {str(e)}")
            embarazos = []
        
        try:
            embarazo_activo = paciente.embarazo_activo
        except Exception as e:
            logger.error(f"Error al obtener embarazo activo: {str(e)}")
            embarazo_activo = None
        
        try:
            consultas_generales = ConsultaGeneral.objects.filter(paciente=paciente.usuario).order_by('-fecha')[:5]
        except Exception as e:
            logger.error(f"Error al obtener consultas generales: {str(e)}")
            consultas_generales = []
        
        try:
            total_consultas = paciente.total_consultas_generales
        except Exception as e:
            logger.error(f"Error al obtener total consultas: {str(e)}")
            total_consultas = 0
        
        try:
            total_embarazos = paciente.total_embarazos
        except Exception as e:
            logger.error(f"Error al obtener total embarazos: {str(e)}")
            total_embarazos = 0
        
        context = {
            'paciente': paciente,
            'embarazos': embarazos,
            'embarazo_activo': embarazo_activo,
            'consultas_generales': consultas_generales,
            'total_consultas': total_consultas,
            'total_embarazos': total_embarazos,
        }
        return render(request, 'admin/ficha_paciente.html', context)
    except Exception as e:
        logger.error(f"Error en ficha_paciente: {str(e)}", exc_info=True)
        messages.error(request, f'Error al cargar la ficha del paciente: {str(e)}')
        return redirect('pacientes_admin')


@login_required
def activar_embarazo_nuevo(request, paciente_id):
    """Activa un nuevo embarazo para la paciente (Doctor prenatal)"""
    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    from pacientes.models import Paciente, Embarazo
    from medicos.models import Medico

    paciente = get_object_or_404(Paciente, id=paciente_id)

    if not es_genero_femenino(paciente.usuario.genero):
        messages.error(request, 'Solo las pacientes registradas como femeninas pueden activar embarazo.')
        return redirect('ficha_paciente', paciente_id=paciente_id)
    
    # Si el género está vacío, asignarlo automáticamente como femenino
    if not paciente.usuario.genero or paciente.usuario.genero.strip() == '':
        paciente.usuario.genero = 'femenino'
        paciente.usuario.save()

    # Verificar que el médico sea prenatal
    try:
        medico_obj = Medico.objects.get(usuario=request.user)
        if not medico_obj.especialidad or medico_obj.especialidad.tipo != 'prenatal':
            messages.error(request, 'Solo médicos prenatales pueden activar embarazos.')
            return redirect('ficha_paciente', paciente_id=paciente_id)
    except Medico.DoesNotExist:
        messages.error(request, 'Debe ser médico prenatal para activar embarazos.')
        return redirect('ficha_paciente', paciente_id=paciente_id)

    if request.method == 'POST':
        try:
            fecha_inicio = request.POST.get('fecha_inicio')
            semanas_inicio = request.POST.get('semanas_gestacion', '0')
            
            if not fecha_inicio:
                messages.error(request, 'La fecha de inicio es obligatoria.')
                return render(request, 'admin/activar_embarazo.html', {'paciente': paciente})
            
            try:
                semanas_inicio = int(semanas_inicio)
            except (ValueError, TypeError):
                semanas_inicio = 0
            
            if semanas_inicio < 0:
                messages.error(request, 'Las semanas de gestación no pueden ser negativas.')
                return render(request, 'admin/activar_embarazo.html', {'paciente': paciente})

            # Crear nuevo embarazo
            embarazo = Embarazo.objects.create(
                paciente=paciente,
                medico_prenatal=request.user,
                fecha_inicio=fecha_inicio,
                semanas_gestacion_inicio=semanas_inicio,
                estado='Activo'
            )

            # Activar embarazo (actualiza paciente)
            embarazo.activar()

            registrar_log(request, 'CREATE', 'Embarazos',
                f'Embarazo activado para {paciente.usuario.get_full_name()}', 'INFO')
            messages.success(request, f'Embarazo activado para {paciente.usuario.first_name}')
            
            # Redirigir según el rol del usuario
            if request.user.rol == 'medico':
                return redirect('pacientes_medico')
            else:
                return redirect('ficha_paciente', paciente_id=paciente_id)
        
        except Exception as e:
            messages.error(request, f'Error al activar embarazo: {str(e)}')
            return render(request, 'admin/activar_embarazo.html', {'paciente': paciente})

    return render(request, 'admin/activar_embarazo.html', {'paciente': paciente})


@login_required
def finalizar_embarazo(request, embarazo_id):
    """Finaliza un embarazo (Doctor prenatal)"""
    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    from pacientes.models import Embarazo

    embarazo = get_object_or_404(Embarazo, id=embarazo_id)

    # Verificar permisos
    if embarazo.medico_prenatal != request.user:
        messages.error(request, 'No tienes permiso para finalizar este embarazo.')
        return redirect('ficha_paciente', paciente_id=embarazo.paciente.id)

    if request.method == 'POST':
        fecha_parto = request.POST.get('fecha_parto')
        semanas_fin = int(request.POST.get('semanas_gestacion_fin', 0))
        tipo_parto = request.POST.get('tipo_parto')
        observaciones = request.POST.get('observaciones', '')

        if not fecha_parto or not tipo_parto:
            messages.error(request, 'Datos incompletos.')
            return render(request, 'admin/finalizar_embarazo.html', {'embarazo': embarazo})

        # Finalizar
        embarazo.finalizar(fecha_parto, semanas_fin, tipo_parto, observaciones)

        registrar_log(request, 'UPDATE', 'Embarazos',
            f'Embarazo finalizado para {embarazo.paciente.usuario.get_full_name()}', 'INFO')
        messages.success(request, 'Embarazo finalizado correctamente.')
        return redirect('ficha_paciente', paciente_id=embarazo.paciente.id)

    return render(request, 'admin/finalizar_embarazo.html', {'embarazo': embarazo})
